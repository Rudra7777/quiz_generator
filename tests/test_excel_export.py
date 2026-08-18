"""
Tests for the combined 'All_Sets' print sheet.

Faculty print one stacked sheet on A4 portrait rather than opening 65 separate
set sheets, so the workbook carries both: the per-set sheets Part 2 reads back,
and one All_Sets sheet holding every set one below another.
"""

import io

from openpyxl import load_workbook

from allocator import QuizStructure, allocate_quizzes, shuffle_all_quizzes
from excel_export import ALL_SETS_SHEET, create_formatted_excel, question_code
from excel_handler import SET_LABEL_RE, load_question_bank, set_label
from response_generator import map_paper_to_bank_questions

from test_answer_pipeline import QUESTION_BANK


HARD, MEDIUM, EASY = 4, 6, 5
QUESTIONS_PER_SET = HARD + MEDIUM + EASY
# Header row + column-header row sit above each set's questions.
BLOCK_ROWS = QUESTIONS_PER_SET + 2


def _papers(num_students: int, seed: int = 7):
    """Build a real papers workbook and hand back the loaded sheet plus its sets."""
    bank = load_question_bank(str(QUESTION_BANK))
    allocation_matrix, usage_counts = allocate_quizzes(
        {
            "hard": bank.get_question_ids_by_difficulty("hard"),
            "medium": bank.get_question_ids_by_difficulty("medium"),
            "easy": bank.get_question_ids_by_difficulty("easy"),
        },
        num_students=num_students,
        quiz_structure=QuizStructure(hard_count=HARD, medium_count=MEDIUM, easy_count=EASY),
        seed=seed,
    )
    shuffled_matrix = shuffle_all_quizzes(allocation_matrix, base_seed=seed)
    data = create_formatted_excel(
        allocation_matrix=allocation_matrix,
        shuffled_matrix=shuffled_matrix,
        usage_counts=usage_counts,
        question_bank=bank,
    )
    return data, shuffled_matrix, bank


def _block_start(set_idx: int) -> int:
    """First row of a set's block on the combined sheet (1-indexed)."""
    return set_idx * BLOCK_ROWS + 1


def test_all_sets_is_the_first_sheet():
    """Faculty should land on the printable sheet when they open the workbook."""
    data, _, _ = _papers(3)
    wb = load_workbook(io.BytesIO(data))

    assert wb.sheetnames[0] == ALL_SETS_SHEET


def test_all_sets_stacks_every_set_in_label_order():
    """Each set gets its own block, in S-01..S-NN order, one below another."""
    data, shuffled_matrix, _ = _papers(4)
    ws = load_workbook(io.BytesIO(data))[ALL_SETS_SHEET]

    labels = [
        ws.cell(row=_block_start(i), column=2).value
        for i in range(len(shuffled_matrix))
    ]

    assert labels == [set_label(i + 1) for i in range(len(shuffled_matrix))]


def test_all_sets_question_rows_match_the_per_set_sheets():
    """The stacked copy of a set must be the same paper as its own sheet."""
    data, shuffled_matrix, _ = _papers(3)
    wb = load_workbook(io.BytesIO(data))
    combined = wb[ALL_SETS_SHEET]

    for set_idx in range(len(shuffled_matrix)):
        per_set = wb[set_label(set_idx + 1)]
        first_row = _block_start(set_idx) + 2
        for q_idx in range(QUESTIONS_PER_SET):
            expected = [per_set.cell(row=q_idx + 3, column=c).value for c in range(1, 9)]
            actual = [combined.cell(row=first_row + q_idx, column=c).value for c in range(1, 9)]
            assert actual == expected, f"{set_label(set_idx + 1)} row {q_idx + 1}"


def test_all_sets_repeats_the_column_headers_for_every_set():
    """A block printed on its own page needs its own Sr/QCd/Question header row."""
    data, shuffled_matrix, _ = _papers(3)
    ws = load_workbook(io.BytesIO(data))[ALL_SETS_SHEET]

    for set_idx in range(len(shuffled_matrix)):
        row = _block_start(set_idx) + 1
        headers = [ws.cell(row=row, column=c).value for c in range(1, 9)]
        assert headers == ['Sr', 'QCd', 'QCd', 'Question', 'A', 'B', 'C', 'D']


def test_all_sets_prints_a4_portrait_fit_to_one_page_wide():
    """The stacked sheet is the print target, so it carries the page setup."""
    data, _, _ = _papers(2)
    ws = load_workbook(io.BytesIO(data))[ALL_SETS_SHEET]

    assert ws.page_setup.orientation == "portrait"
    assert str(ws.page_setup.paperSize) == str(ws.PAPERSIZE_A4)
    assert ws.page_setup.fitToWidth == 1
    assert ws.page_setup.fitToHeight == 0
    assert ws.sheet_properties.pageSetUpPr.fitToPage is True


def test_all_sets_visible_columns_fit_a4_portrait_width():
    """Columns must total the ~92 units A4 portrait prints, or text is shrunk."""
    data, _, _ = _papers(2)
    ws = load_workbook(io.BytesIO(data))[ALL_SETS_SHEET]

    assert ws.column_dimensions['B'].hidden is True
    visible = sum(
        ws.column_dimensions[ch].width for ch in 'ACDEFGH'
    )
    assert visible == 92


def test_all_sets_starts_each_set_on_a_new_page():
    """No set may begin halfway down a page - papers have to be separable."""
    data, shuffled_matrix, _ = _papers(4)
    ws = load_workbook(io.BytesIO(data))[ALL_SETS_SHEET]

    breaks = sorted(brk.id for brk in ws.row_breaks.brk)

    assert breaks == [_block_start(i) - 1 for i in range(1, len(shuffled_matrix))]


def test_all_sets_uses_the_printed_question_code():
    """Column C carries the 'Q- 27' form students copy into the Google Form."""
    data, shuffled_matrix, bank = _papers(2)
    ws = load_workbook(io.BytesIO(data))[ALL_SETS_SHEET]

    first_qid = shuffled_matrix[0][0]
    expected = question_code(bank.get_by_id(first_qid).question_no)

    assert ws.cell(row=_block_start(0) + 2, column=3).value == expected


def test_per_set_sheets_survive_alongside_the_combined_sheet():
    """Part 2 finds sets by sheet name - All_Sets must not join or disturb them."""
    data, shuffled_matrix, _ = _papers(3)
    wb = load_workbook(io.BytesIO(data))

    set_sheets = [name for name in wb.sheetnames if SET_LABEL_RE.match(name)]

    assert set_sheets == [set_label(i + 1) for i in range(len(shuffled_matrix))]
    assert not SET_LABEL_RE.match(ALL_SETS_SHEET)


def test_part_two_reader_still_maps_the_new_workbook(tmp_path):
    """End-to-end guard: answer_checker's reader keeps working on a regenerated file."""
    data, shuffled_matrix, bank = _papers(3)
    path = tmp_path / "question_papers.xlsx"
    path.write_bytes(data)

    set_map = map_paper_to_bank_questions(str(path), bank)

    assert sorted(set_map) == [set_label(i + 1) for i in range(len(shuffled_matrix))]
    for set_idx, label in enumerate(sorted(set_map)):
        expected = [bank.get_by_id(qid).question_no for qid in shuffled_matrix[set_idx]]
        assert set_map[label] == expected
