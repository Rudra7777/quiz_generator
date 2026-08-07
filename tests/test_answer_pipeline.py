from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from answer_checker import (
    check_all_responses,
    generate_scoring_report,
    load_response_sheet,
    normalize_response_sheet,
)
from excel_handler import load_question_bank
from response_generator import (
    ROLL_COL,
    SET_COL,
    TIMESTAMP_COL,
    answer_column,
    generate_responses,
    map_paper_to_bank_questions,
)


ROOT = Path(__file__).resolve().parents[1]
QUESTION_BANK = ROOT / "input" / "question_bank_72.xlsx"
QUESTION_PAPERS = ROOT / "output" / "question_papers.xlsx"
RESPONSES = ROOT / "output" / "student_responses.xlsx"


def _responses(num_students: int, seed: int) -> pd.DataFrame:
    """Dummy responses in the canonical shape check_all_responses expects."""
    question_bank = load_question_bank(str(QUESTION_BANK))
    return normalize_response_sheet(
        generate_responses(
            question_papers_path=str(QUESTION_PAPERS),
            question_bank=question_bank,
            num_students=num_students,
            seed=seed,
        )
    )


def test_regression_existing_response_sheet():
    """Regression: current fixture responses should keep known metrics."""
    question_bank = load_question_bank(str(QUESTION_BANK))
    response_df = load_response_sheet(str(RESPONSES))

    report = check_all_responses(
        response_df=response_df,
        question_papers_path=str(QUESTION_PAPERS),
        question_bank=question_bank,
    )

    assert len(report.student_reports) == 70
    assert report.avg_score == 10.54
    assert report.median_score == 10.5
    assert report.pass_rate == 100.0
    assert report.pass_count == 70
    assert len(report.validation_issues) == 0
    assert report.grade_distribution() == {
        "14/15": 3,
        "13/15": 10,
        "12/15": 13,
        "11/15": 9,
        "10/15": 11,
        "9/15": 14,
        "8/15": 7,
        "7/15": 1,
        "6/15": 2,
    }


def test_seeded_generation_is_deterministic():
    """Same seed + same inputs should produce identical response sheets."""
    question_bank = load_question_bank(str(QUESTION_BANK))

    df1 = _responses(20, seed=2026)
    df2 = _responses(20, seed=2026)

    pd.testing.assert_frame_equal(df1, df2, check_dtype=False)

    report = check_all_responses(
        response_df=df1,
        question_papers_path=str(QUESTION_PAPERS),
        question_bank=question_bank,
    )
    assert len(report.validation_issues) == 0
    assert all(r.assigned == 15 for r in report.student_reports)
    assert all(r.attempted == r.assigned for r in report.student_reports)
    assert all(r.unanswered == 0 for r in report.student_reports)
    assert all(r.correct + r.wrong == r.assigned for r in report.student_reports)


def test_generated_responses_carry_student_identity():
    """The dummy sheet must look like a real Google Forms export."""
    raw = generate_responses(
        question_papers_path=str(QUESTION_PAPERS),
        question_bank=load_question_bank(str(QUESTION_BANK)),
        num_students=3,
        seed=11,
    )

    assert list(raw.columns[:5]) == [
        "Timestamp",
        "Email address",
        "Full Name",
        "Roll Number",
        "Question Set",
    ]
    assert raw.columns[5] == answer_column(1)
    assert raw[SET_COL].tolist() == ["S-01", "S-02", "S-03"]
    assert raw[ROLL_COL].tolist() == ["R001", "R002", "R003"]


def test_answer_column_headers_are_read_tolerantly():
    """Google's header, and the shorter forms people hand-edit to, all mean Q7."""
    for header in ("Q - 07 [Answer]", "Q-07", "Q07", "Q7"):
        df = pd.DataFrame(
            [{
                TIMESTAMP_COL: "2026-02-16 16:00:00",
                "Email address": "a@example.com",
                "Full Name": "A",
                ROLL_COL: "R001",
                SET_COL: "S-01",
                header: "B",
            }]
        )
        normalized = normalize_response_sheet(df)
        assert "Q7" in normalized.columns, f"{header!r} did not normalize to Q7"
        assert normalized.loc[0, "Q7"] == "B"


def test_latest_submission_wins_and_the_dropped_one_is_reported():
    """A student who submits twice is scored on their later attempt."""
    question_bank = load_question_bank(str(QUESTION_BANK))
    response_df = _responses(5, seed=3)

    set_map = map_paper_to_bank_questions(str(QUESTION_PAPERS), question_bank)
    assigned = set_map["S-01"]
    qno_to_answer = {q.question_no: q.answer for q in question_bank.get_all()}

    # Row 0's resubmission gets every assigned question right, an hour later.
    resubmission = response_df.loc[0].copy()
    resubmission[TIMESTAMP_COL] = response_df.loc[0, TIMESTAMP_COL] + pd.Timedelta(hours=1)
    for q_no in assigned:
        resubmission[f"Q{q_no}"] = qno_to_answer[q_no]

    with_duplicate = pd.concat(
        [response_df, resubmission.to_frame().T], ignore_index=True
    )

    report = check_all_responses(
        response_df=with_duplicate,
        question_papers_path=str(QUESTION_PAPERS),
        question_bank=question_bank,
    )

    assert len(report.student_reports) == 5, "duplicate should not add a student"
    assert len(report.superseded) == 1
    assert report.superseded[0].roll_number == "R001"

    scored = next(r for r in report.student_reports if r.roll_number == "R001")
    assert scored.correct == scored.assigned, "the later, all-correct attempt should win"


def test_validation_flags_extra_answer_on_unassigned_question(tmp_path: Path):
    """If student answers outside assigned set, validation must report it."""
    question_bank = load_question_bank(str(QUESTION_BANK))
    response_df = _responses(5, seed=7)

    set_map = map_paper_to_bank_questions(str(QUESTION_PAPERS), question_bank)
    assigned_qnos = set(set_map["S-01"])
    total_qnos = set(range(1, len(question_bank.get_all()) + 1))
    extra_qno = min(total_qnos - assigned_qnos)
    response_df.loc[0, f"Q{extra_qno}"] = "A"

    report = check_all_responses(
        response_df=response_df,
        question_papers_path=str(QUESTION_PAPERS),
        question_bank=question_bank,
    )

    assert len(report.validation_issues) >= 1
    first = report.validation_issues[0]
    assert first.student_index == 0
    assert first.validation.extra_count >= 1
    assert extra_qno in first.validation.extra_questions

    output_path = tmp_path / "scoring_report_with_issue.xlsx"
    saved = generate_scoring_report(
        report,
        str(output_path),
        question_papers_path=str(QUESTION_PAPERS),
        question_bank=question_bank,
    )
    assert Path(saved).exists()

    xl = pd.ExcelFile(saved)
    assert "Responses_Review" in xl.sheet_names

    wb = load_workbook(saved)
    ws = wb["Responses_Review"]
    has_colored = False
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if cell.fill and cell.fill.fill_type == "solid":
                color = (cell.fill.start_color.rgb or "").upper()
                if color.endswith("C6EFCE") or color.endswith("FFC7CE"):
                    has_colored = True
                    break
        if has_colored:
            break
    assert has_colored

    validation_df = pd.read_excel(saved, sheet_name="Validation")
    assert "Issue" in validation_df.columns
    assert validation_df["Issue"].str.contains("outside their set").any()


def test_summary_has_charts_without_disturbing_the_metrics_table(tmp_path: Path):
    """Charts go in D onward; the faculty's Metric/Value table stays in A-B."""
    question_bank = load_question_bank(str(QUESTION_BANK))
    response_df = _responses(30, seed=9)

    report = check_all_responses(
        response_df=response_df,
        question_papers_path=str(QUESTION_PAPERS),
        question_bank=question_bank,
    )
    saved = generate_scoring_report(
        report,
        str(tmp_path / "report.xlsx"),
        question_papers_path=str(QUESTION_PAPERS),
        question_bank=question_bank,
    )

    ws = load_workbook(saved)["Summary"]

    assert ws["A1"].value == "Metric"
    assert ws["B1"].value == "Value"
    assert ws["A2"].value == "Total Students"
    assert ws["B2"].value == 30

    assert [type(c).__name__ for c in ws._charts] == ["BarChart"]

    # openpyxl hides axes by default, which drops the tick numbers.
    chart = ws._charts[0]
    assert chart.x_axis.delete is False
    assert chart.y_axis.delete is False

    # Histogram covers every mark from 0 to max, gaps included, and the bars
    # must account for every student exactly once.
    max_marks = max(r.assigned for r in report.student_reports)
    assert ws["D1"].value == "Marks"
    marks = [ws.cell(2 + i, 4).value for i in range(max_marks + 1)]
    assert marks == list(range(max_marks + 1))

    bars = [ws.cell(2 + i, 5).value for i in range(max_marks + 1)]
    assert sum(bars) == 30

    # The fitted curve should peak at the mark nearest the mean.
    curve = [ws.cell(2 + i, 6).value for i in range(max_marks + 1)]
    assert curve.index(max(curve)) == round(report.avg_score)


def test_out_of_set_answers_are_attempted_but_never_earn_marks(tmp_path: Path):
    """A correct answer to somebody else's question must score nothing."""
    question_bank = load_question_bank(str(QUESTION_BANK))
    response_df = _responses(5, seed=21)

    set_map = map_paper_to_bank_questions(str(QUESTION_PAPERS), question_bank)
    assigned = set(set_map["S-01"])
    key = {q.question_no: q.answer for q in question_bank.get_all()}

    # Give student 1 two extra questions and let them get BOTH right.
    extras = [q for q in sorted(key) if q not in assigned][:2]
    for q_no in extras:
        response_df.loc[0, f"Q{q_no}"] = key[q_no]

    report = check_all_responses(
        response_df=response_df,
        question_papers_path=str(QUESTION_PAPERS),
        question_bank=question_bank,
    )
    student = report.student_reports[0]

    assert student.assigned == 15
    assert student.attempted == 17, "extras count towards attempted"
    assert student.correct <= 15, "extras must not inflate the mark"
    assert student.validation.extra_count == 2

    saved = generate_scoring_report(
        report,
        str(tmp_path / "report.xlsx"),
        question_papers_path=str(QUESTION_PAPERS),
        question_bank=question_bank,
    )
    ws = load_workbook(saved)["Faculty_Report"]

    num_questions = len(question_bank.get_all())
    first_col, last_col = 8, 7 + num_questions
    mask_first = last_col + 2

    key_row = [ws.cell(2, c).value for c in range(first_col, last_col + 1)]
    answers = [ws.cell(4, c).value for c in range(first_col, last_col + 1)]
    mask = [ws.cell(4, c).value for c in range(mask_first, mask_first + num_questions)]

    # Evaluate the sheet's own formulas the way Excel would.
    counta = sum(1 for v in answers if v not in (None, ""))
    ansc = sum(m for k, v, m in zip(key_row, answers, mask) if k == v)

    assert counta == student.attempted == 17
    assert ansc == student.correct, "AnsC must exclude out-of-set answers"

    # The two extras are answered, correct against the key, and masked out.
    for q_no in extras:
        col = first_col + q_no - 1
        assert ws.cell(4, col).value == key[q_no]
        assert ws.cell(4, mask_first + q_no - 1).value == 0


def test_faculty_report_layout(tmp_path: Path):
    """The faculty sheet: key strip on top, live Count/AnsC, one row per student."""
    question_bank = load_question_bank(str(QUESTION_BANK))
    response_df = _responses(4, seed=5)

    report = check_all_responses(
        response_df=response_df,
        question_papers_path=str(QUESTION_PAPERS),
        question_bank=question_bank,
    )

    saved = generate_scoring_report(
        report,
        str(tmp_path / "report.xlsx"),
        question_papers_path=str(QUESTION_PAPERS),
        question_bank=question_bank,
    )

    ws = load_workbook(saved)["Faculty_Report"]
    num_questions = len(question_bank.get_all())
    first_col, last_col = 8, 7 + num_questions

    assert ws.max_row == 3 + 4  # header block + one row per student

    # Row 2 is the full answer key, so no blank-vs-blank pair can inflate AnsC.
    key = [ws.cell(2, c).value for c in range(first_col, last_col + 1)]
    assert all(k in ("A", "B", "C", "D") for k in key)
    assert ws.cell(1, first_col).value == "Q01"

    headers = [ws.cell(3, c).value for c in range(1, 8)]
    assert headers == [
        "Timestamp",
        "Email address",
        "Full Name",
        "Roll Number",
        "Question Set",
        "Count",
        "AnsC",
    ]

    # Formulas must span exactly the question columns — see docs/adr/0002 — and
    # AnsC must be gated by the hidden assigned-question helper block.
    last = get_column_letter(last_col)
    mask_first = get_column_letter(last_col + 2)
    mask_last = get_column_letter(last_col + 1 + num_questions)
    assert ws["F4"].value == f"=COUNTA(H4:{last}4)"
    assert ws["G4"].value == (
        f"=SUMPRODUCT(--(H$2:{last}$2=H4:{last}4),{mask_first}4:{mask_last}4)"
    )
    assert ws.column_dimensions[mask_first].hidden

    # AnsC is the true mark: it must agree with the Scores sheet.
    scores = pd.read_excel(saved, sheet_name="Scores")
    for offset, expected in enumerate(scores["Correct"]):
        student = [ws.cell(4 + offset, c).value for c in range(first_col, last_col + 1)]
        assert sum(1 for k, v in zip(key, student) if k == v) == expected


def test_faculty_report_colour_codes_answers_against_the_key(tmp_path: Path):
    """Text says right/wrong against the key; background says whose question it was."""
    question_bank = load_question_bank(str(QUESTION_BANK))
    response_df = _responses(4, seed=13)

    report = check_all_responses(
        response_df=response_df,
        question_papers_path=str(QUESTION_PAPERS),
        question_bank=question_bank,
    )
    saved = generate_scoring_report(
        report,
        str(tmp_path / "report.xlsx"),
        question_papers_path=str(QUESTION_PAPERS),
        question_bank=question_bank,
    )

    num_questions = len(question_bank.get_all())
    ws = load_workbook(saved)["Faculty_Report"]
    last = get_column_letter(7 + num_questions)
    mask = get_column_letter(9 + num_questions)

    ranges = list(ws.conditional_formatting)
    assert len(ranges) == 1
    assert str(ranges[0].sqref) == f"H4:{last}7"

    styles = {rule.formula[0]: rule.dxf for rule in ranges[0].rules}

    # Text colour is the answer's verdict, and says nothing about whose question
    # it was — an out-of-set answer that matches the key still reads blue.
    correct = styles['AND(H4<>"",H4=H$2)'].font
    assert correct.color.rgb == "FF0000FF" and correct.b and not correct.i

    wrong = styles['AND(H4<>"",H4<>H$2)'].font
    assert wrong.color.rgb == "FFFF0000" and wrong.b and wrong.i

    # Background is allocation. Green covers the whole of a student's own set,
    # answered or not, so the questions they skipped show as empty green cells.
    assert styles[f"{mask}4=1"].fill.bgColor.rgb == "FFC6EFCE"

    # Yellow only where they strayed outside their set — an unanswered question
    # that was never theirs stays unfilled.
    assert styles[f'AND({mask}4=0,H4<>"")'].fill.bgColor.rgb == "FFFFF2CC"

    # Correct/wrong rules key off row 2 — the answer key — so that editing a key
    # letter recolours the column.
    assert all("$2" in f for f in styles if "H4=H$2" in f or "H4<>H$2" in f)
