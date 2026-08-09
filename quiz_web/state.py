"""
Reflex state for the Quiz Generator UI.

Two states mirror the two pages:
- GenerateState: upload a question bank and produce randomized question papers.
- EvaluateState: generate dummy responses, then validate and score submissions.

All heavy lifting is delegated to the existing engine modules; raw file bytes are
held in backend-only vars (leading underscore) so they are never shipped to the client.
"""

import io
import os
import secrets
import tempfile

import reflex as rx
from openpyxl import load_workbook

from allocator import QuizStructure, allocate_quizzes, shuffle_all_quizzes
from excel_handler import load_question_bank, SET_LABEL_RE
from response_generator import generate_responses
from answer_checker import (
    load_response_sheet,
    check_all_responses,
    generate_scoring_report,
)
from excel_export import (
    create_formatted_excel,
    _make_excel_bytes_from_dataframe,
    _load_question_bank_from_question_papers,
)


def _to_int(value, default: int = 0) -> int:
    """Best-effort parse of a form string into an int."""
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return default


def _write_temp(data: bytes) -> str:
    """Persist raw bytes to a temporary .xlsx path and return it."""
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx", prefix="quiz_web_") as tmp:
        tmp.write(data)
        return tmp.name


class GenerateState(rx.State):
    """Part 1 — generate question papers."""

    # Backend-only (never synced to the browser)
    _bank_bytes: bytes = b""
    _papers_bytes: bytes = b""

    # Uploaded bank
    bank_uploaded: bool = False
    bank_filename: str = ""
    hard_avail: int = 0
    medium_avail: int = 0
    easy_avail: int = 0

    # Configuration
    num_students: int = 50
    total_questions: int = 15
    mode: str = "Absolute"  # "Absolute" | "Percentage"

    hard_count: int = 4
    medium_count: int = 6
    easy_count: int = 5

    hard_pct: int = 27
    medium_pct: int = 40
    easy_pct: int = 33

    use_fixed_seed: bool = False
    seed: int = 42

    # Output / feedback
    status: str = ""
    error: str = ""
    papers_ready: bool = False
    last_seed: int = 0
    generated_count: int = 0

    # ── Derived values ────────────────────────────────────────────────────
    @rx.var
    def total_avail(self) -> int:
        return self.hard_avail + self.medium_avail + self.easy_avail

    @rx.var
    def calc_hard(self) -> int:
        if self.mode == "Percentage":
            return round(self.total_questions * self.hard_pct / 100)
        return self.hard_count

    @rx.var
    def calc_medium(self) -> int:
        if self.mode == "Percentage":
            return round(self.total_questions * self.medium_pct / 100)
        return self.medium_count

    @rx.var
    def calc_easy(self) -> int:
        if self.mode == "Percentage":
            return self.total_questions - self.calc_hard - self.calc_medium
        return self.easy_count

    @rx.var
    def total_selected(self) -> int:
        return self.calc_hard + self.calc_medium + self.calc_easy

    @rx.var
    def pct_sum(self) -> int:
        return self.hard_pct + self.medium_pct + self.easy_pct

    @rx.var
    def size_ok(self) -> bool:
        return self.total_selected == self.total_questions

    @rx.var
    def avail_ok(self) -> bool:
        return (
            self.calc_hard <= self.hard_avail
            and self.calc_medium <= self.medium_avail
            and self.calc_easy <= self.easy_avail
            and self.calc_hard >= 0
            and self.calc_medium >= 0
            and self.calc_easy >= 0
        )

    @rx.var
    def can_generate(self) -> bool:
        return self.bank_uploaded and self.size_ok and self.avail_ok

    # ── Simple field setters ──────────────────────────────────────────────
    @rx.event
    def set_num_students(self, value: str):
        self.num_students = _to_int(value, self.num_students)

    @rx.event
    def set_total_questions(self, value: str):
        self.total_questions = _to_int(value, self.total_questions)

    @rx.event
    def set_mode(self, value: str):
        self.mode = value

    @rx.event
    def set_hard_count(self, value: str):
        self.hard_count = _to_int(value)

    @rx.event
    def set_medium_count(self, value: str):
        self.medium_count = _to_int(value)

    @rx.event
    def set_easy_count(self, value: str):
        self.easy_count = _to_int(value)

    @rx.event
    def set_hard_pct(self, value: list[int | float]):
        self.hard_pct = int(value[0])

    @rx.event
    def set_medium_pct(self, value: list[int | float]):
        self.medium_pct = int(value[0])

    @rx.event
    def set_easy_pct(self, value: list[int | float]):
        self.easy_pct = int(value[0])

    @rx.event
    def toggle_fixed_seed(self, value: bool):
        self.use_fixed_seed = value

    @rx.event
    def set_seed(self, value: str):
        self.seed = _to_int(value, self.seed)

    # ── Actions ───────────────────────────────────────────────────────────
    @rx.event
    async def handle_bank_upload(self, files: list[rx.UploadFile]):
        self.status = ""
        self.error = ""
        self.papers_ready = False
        if not files:
            return
        data = await files[0].read()
        path = _write_temp(data)
        try:
            bank = load_question_bank(path)
            counts = bank.count_by_difficulty()
            self.hard_avail = counts.get("hard", 0)
            self.medium_avail = counts.get("medium", 0)
            self.easy_avail = counts.get("easy", 0)
            self._bank_bytes = data
            self.bank_uploaded = True
            self.bank_filename = files[0].filename or ""
            # Seed sensible defaults from what's available.
            self.hard_count = min(4, self.hard_avail)
            self.medium_count = min(6, self.medium_avail)
            self.easy_count = min(5, self.easy_avail)
            self.status = f"Loaded {self.total_avail} questions."
        except Exception as exc:  # noqa: BLE001 - surface any load error to the user
            self.bank_uploaded = False
            self.bank_filename = ""
            self.error = f"Could not load file: {exc}"
        finally:
            if os.path.exists(path):
                os.remove(path)

    @rx.event
    def generate_papers(self):
        self.status = ""
        self.error = ""
        self.papers_ready = False
        if not self.can_generate:
            self.error = "Fix the configuration before generating."
            return
        if not self._bank_bytes:
            self.bank_uploaded = False
            self.error = "The uploaded question bank expired. Upload it again."
            return
        path = _write_temp(self._bank_bytes)
        try:
            bank = load_question_bank(path)
            quiz_structure = QuizStructure(
                hard_count=self.calc_hard,
                medium_count=self.calc_medium,
                easy_count=self.calc_easy,
            )
            q_ids_by_diff = {
                "hard": bank.get_question_ids_by_difficulty("hard"),
                "medium": bank.get_question_ids_by_difficulty("medium"),
                "easy": bank.get_question_ids_by_difficulty("easy"),
            }
            run_seed = int(self.seed) if self.use_fixed_seed else secrets.randbelow(2_147_483_647)
            allocation_matrix, usage_counts = allocate_quizzes(
                q_ids_by_diff,
                num_students=self.num_students,
                quiz_structure=quiz_structure,
                seed=run_seed,
            )
            shuffled_matrix = shuffle_all_quizzes(allocation_matrix, base_seed=run_seed)
            self._papers_bytes = create_formatted_excel(
                allocation_matrix=allocation_matrix,
                shuffled_matrix=shuffled_matrix,
                usage_counts=usage_counts,
                question_bank=bank,
            )
            self.papers_ready = True
            self.last_seed = run_seed
            self.generated_count = self.num_students
            self.status = f"Generated {self.num_students} question papers."
            # Deliver in this same event so the bytes never have to survive a round-trip
            # through the session store — see the note on _build_scoring_report.
            return rx.download(data=self._papers_bytes, filename="question_papers.xlsx")
        except Exception as exc:  # noqa: BLE001
            self.error = f"Generation failed: {exc}"
        finally:
            if os.path.exists(path):
                os.remove(path)

    @rx.event
    def download_papers(self):
        if not self._papers_bytes:
            self.error = "The generated papers expired. Generate them again."
            return
        return rx.download(data=self._papers_bytes, filename="question_papers.xlsx")


class EvaluateState(rx.State):
    """Part 2 — generate dummy responses and score submissions."""

    # Backend-only bytes
    _gen_papers_bytes: bytes = b""
    _generated_responses_bytes: bytes = b""
    _chk_papers_bytes: bytes = b""
    _chk_responses_bytes: bytes = b""

    # Generate dummy responses
    gen_papers_uploaded: bool = False
    gen_papers_filename: str = ""
    gen_students: int = 70
    correct_rate: int = 70
    wrong_rate: int = 20
    gen_use_seed: bool = False
    gen_seed: int = 42
    gen_status: str = ""
    gen_error: str = ""
    responses_ready: bool = False

    # Check & score
    chk_papers_uploaded: bool = False
    chk_papers_filename: str = ""
    chk_responses_uploaded: bool = False
    chk_responses_filename: str = ""
    pass_threshold: int = 6
    score_status: str = ""
    score_error: str = ""
    report_ready: bool = False
    avg_score: float = 0.0
    median_score: float = 0.0
    pass_rate: float = 0.0
    max_marks: int = 0
    validation_count: int = 0
    total_students: int = 0

    @rx.var
    def extra_rate(self) -> int:
        return 100 - self.correct_rate - self.wrong_rate

    @rx.var
    def rates_ok(self) -> bool:
        return self.extra_rate >= 0

    # ── Setters ───────────────────────────────────────────────────────────
    @rx.event
    def set_gen_students(self, value: str):
        self.gen_students = _to_int(value, self.gen_students)

    @rx.event
    def set_correct_rate(self, value: list[int | float]):
        self.correct_rate = int(value[0])

    @rx.event
    def set_wrong_rate(self, value: list[int | float]):
        self.wrong_rate = int(value[0])

    @rx.event
    def toggle_gen_seed(self, value: bool):
        self.gen_use_seed = value

    @rx.event
    def set_gen_seed(self, value: str):
        self.gen_seed = _to_int(value, self.gen_seed)

    @rx.event
    def set_pass_threshold(self, value: str):
        self.pass_threshold = _to_int(value, self.pass_threshold)

    # ── Uploads ───────────────────────────────────────────────────────────
    @rx.event
    async def handle_gen_papers_upload(self, files: list[rx.UploadFile]):
        self.gen_status = ""
        self.gen_error = ""
        self.responses_ready = False
        if not files:
            return
        data = await files[0].read()
        self._gen_papers_bytes = data
        self.gen_papers_uploaded = True
        self.gen_papers_filename = files[0].filename or ""

        # Auto-fill student count from the number of set sheets in the file.
        try:
            wb = load_workbook(io.BytesIO(data), read_only=True)
            set_count = sum(1 for name in wb.sheetnames if SET_LABEL_RE.match(name))
            wb.close()
            if set_count:
                self.gen_students = set_count
        except Exception:  # noqa: BLE001 - keep existing student count if unreadable
            pass

        self.gen_status = "Question papers loaded."

    @rx.event
    async def handle_chk_papers_upload(self, files: list[rx.UploadFile]):
        self.score_error = ""
        if not files:
            return
        self._chk_papers_bytes = await files[0].read()
        self.chk_papers_uploaded = True
        self.chk_papers_filename = files[0].filename or ""

    @rx.event
    async def handle_chk_responses_upload(self, files: list[rx.UploadFile]):
        self.score_error = ""
        if not files:
            return
        self._chk_responses_bytes = await files[0].read()
        self.chk_responses_uploaded = True
        self.chk_responses_filename = files[0].filename or ""

    # ── Actions ───────────────────────────────────────────────────────────
    @rx.event
    def generate_responses_action(self):
        self.gen_status = ""
        self.gen_error = ""
        self.responses_ready = False
        if not self.gen_papers_uploaded:
            self.gen_error = "Upload a question papers file first."
            return
        if not self.rates_ok:
            self.gen_error = "Correct % + Wrong % cannot exceed 100."
            return
        path = _write_temp(self._gen_papers_bytes)
        try:
            bank = _load_question_bank_from_question_papers(path)
            response_df = generate_responses(
                question_papers_path=path,
                question_bank=bank,
                num_students=self.gen_students,
                correct_rate=self.correct_rate / 100.0,
                wrong_rate=self.wrong_rate / 100.0,
                blank_rate=0.0,
                seed=self.gen_seed if self.gen_use_seed else None,
            )
            responses_bytes = _make_excel_bytes_from_dataframe(response_df, "Responses")
            self._generated_responses_bytes = responses_bytes
            self.responses_ready = True
            self.gen_status = f"Generated dummy responses for {len(response_df)} students."
            # Deliver in this same event so the exact generated data doesn't depend
            # on the session store keeping the blob across events.
            return rx.download(data=responses_bytes, filename="student_responses.xlsx")
        except Exception as exc:  # noqa: BLE001
            self.gen_error = f"Could not generate responses: {exc}"
        finally:
            if os.path.exists(path):
                os.remove(path)

    @rx.event
    def download_responses(self):
        if not self._generated_responses_bytes:
            self.gen_error = "The generated responses expired. Generate them again."
            return
        return rx.download(
            data=self._generated_responses_bytes, filename="student_responses.xlsx"
        )

    def _build_scoring_report(self):
        """Score the uploaded responses and return (report, report_bytes).

        Pure computation from the two uploaded blobs — no state writes. Scoring is
        deterministic, so this can be re-run on a later download click and produce
        the identical file. Keeping the report out of persisted state avoids relying
        on Reflex Cloud's Redis session store holding a large blob across events
        (under memory pressure it may evict it, yielding a 0-byte download).
        """
        qp_path = _write_temp(self._chk_papers_bytes)
        resp_path = _write_temp(self._chk_responses_bytes)
        try:
            bank = _load_question_bank_from_question_papers(qp_path)
            response_df = load_response_sheet(resp_path)
            report = check_all_responses(
                response_df=response_df,
                question_papers_path=qp_path,
                question_bank=bank,
                pass_threshold=float(self.pass_threshold),
            )
            report_buffer = io.BytesIO()
            generate_scoring_report(
                report,
                report_buffer,
                question_papers_path=qp_path,
                question_bank=bank,
            )
            return report, report_buffer.getvalue()
        finally:
            for p in (qp_path, resp_path):
                if p and os.path.exists(p):
                    os.remove(p)

    @rx.event
    def check_score_action(self):
        self.score_status = ""
        self.score_error = ""
        self.report_ready = False
        if not self.chk_papers_uploaded:
            self.score_error = "Upload the question papers file."
            return
        if not self.chk_responses_uploaded:
            self.score_error = "Upload the student responses file."
            return
        try:
            report, report_bytes = self._build_scoring_report()
            self.avg_score = report.avg_score
            self.median_score = report.median_score
            self.pass_rate = report.pass_rate
            self.max_marks = (
                report.student_reports[0].assigned if report.student_reports else 0
            )
            self.validation_count = len(report.validation_issues)
            self.total_students = len(report.student_reports)
            self.report_ready = True
            self.score_status = "Scoring completed."
            # Deliver the file in this same event so the bytes never have to survive
            # a round-trip through the session store.
            return rx.download(data=report_bytes, filename="scoring_report.xlsx")
        except Exception as exc:  # noqa: BLE001
            self.score_error = f"Scoring failed: {exc}"

    @rx.event
    def download_report(self):
        try:
            _, report_bytes = self._build_scoring_report()
            return rx.download(data=report_bytes, filename="scoring_report.xlsx")
        except Exception as exc:  # noqa: BLE001
            self.score_error = f"Scoring failed: {exc}"
