"""
Answer Checker Module

Validates student responses and computes scoring reports.
"""

from __future__ import annotations

from collections import Counter
from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, List, Optional
import math
import re
import statistics

import pandas as pd
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.formatting.rule import Rule
from openpyxl.styles import Font, PatternFill
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.utils import get_column_letter

from excel_handler import FullQuestionBank, parse_set_label
from response_generator import (
    ANSWER_COL_RE,
    EMAIL_COL,
    IDENTITY_COLS,
    NAME_COL,
    ROLL_COL,
    SET_COL,
    TIMESTAMP_COL,
    map_paper_to_bank_questions,
)


VALID_OPTIONS = {"A", "B", "C", "D"}
QUESTION_COL_RE = re.compile(r"^Q(\d+)$")

# Columns the response sheet cannot be scored without: the join key to the answer
# key, plus what the "latest submission wins" rule needs to pick a winner.
REQUIRED_RESPONSE_COLS = (SET_COL, ROLL_COL, TIMESTAMP_COL)


@dataclass
class ValidationResult:
    """Validation details for one student's response row."""

    set_no: str
    extra_questions: List[int]

    @property
    def extra_count(self) -> int:
        return len(self.extra_questions)


@dataclass
class StudentReport:
    """Scoring details for one student."""

    student_index: int
    validation: ValidationResult
    assigned: int
    attempted: int
    correct: int
    wrong: int
    unanswered: int
    roll_number: str = ""
    full_name: str = ""
    email: str = ""

    @property
    def set_no(self) -> str:
        return self.validation.set_no


@dataclass
class SupersededSubmission:
    """A submission dropped because the same Roll Number submitted again later."""

    roll_number: str
    set_no: str
    submitted_at: object
    kept_at: object


@dataclass
class ScoringReport:
    """Aggregated scoring report for all students."""

    student_reports: List[StudentReport]
    validation_issues: List[StudentReport]
    avg_score: float
    median_score: float
    pass_count: int
    pass_rate: float
    pass_threshold: float = 6.0
    superseded: List[SupersededSubmission] = field(default_factory=list)
    scored_df: Optional[pd.DataFrame] = None

    def grade_distribution(self) -> Dict[str, int]:
        """Return distribution by obtained marks (Correct / Max Marks)."""
        buckets: Dict[str, int] = {}
        for report in self.student_reports:
            key = f"{report.correct}/{report.assigned}"
            buckets[key] = buckets.get(key, 0) + 1
        # Sort by marks descending for readability.
        buckets = dict(sorted(buckets.items(), key=lambda kv: int(kv[0].split("/")[0]), reverse=True))
        return buckets


def _cell_text(value: object) -> str:
    """Cell value as a trimmed string, with pandas' blanks flattened to ''."""
    if value is None or pd.isna(value):
        return ""
    return str(value).strip()


def _normalize_answer(value: object) -> Optional[str]:
    """Normalize answer value to A/B/C/D, or None for blank."""
    if pd.isna(value):
        return None
    normalized = str(value).strip().upper()
    if not normalized:
        return None
    return normalized


def _extract_answered_questions(row: pd.Series) -> Dict[int, str]:
    """Extract non-blank answered question numbers from a response row."""
    answered: Dict[int, str] = {}
    for col, value in row.items():
        match = QUESTION_COL_RE.match(str(col))
        if not match:
            continue
        answer = _normalize_answer(value)
        if answer is None:
            continue
        answered[int(match.group(1))] = answer
    return answered


def normalize_response_sheet(df: pd.DataFrame) -> pd.DataFrame:
    """
    Put a Google Forms export into canonical form.

    Identity columns keep their exported names; answer columns collapse from
    "Q - 01 [Answer]" to "Q1" so they key directly off the question number.
    Rows come back ordered by set, which is the order the faculty report wants.
    """
    out = df.copy()
    out.columns = [str(c).strip() for c in out.columns]

    # Match identity columns tolerantly — trailing spaces and casing vary between
    # a hand-built Form and a real export.
    canonical_identity = {c.lower(): c for c in IDENTITY_COLS}
    renames = {}
    for col in out.columns:
        target = canonical_identity.get(col.lower())
        if target is not None:
            renames[col] = target
            continue
        match = ANSWER_COL_RE.match(col)
        if match:
            renames[col] = f"Q{int(match.group(1))}"
    out = out.rename(columns=renames)

    missing = [c for c in REQUIRED_RESPONSE_COLS if c not in out.columns]
    if missing:
        raise ValueError(
            f"Responses sheet is missing required column(s): {', '.join(missing)}. "
            f"Expected a Google Forms export with columns: {', '.join(IDENTITY_COLS)}, "
            "then one 'Q - NN [Answer]' column per question."
        )

    for optional in (EMAIL_COL, NAME_COL):
        if optional not in out.columns:
            out[optional] = ""

    out[TIMESTAMP_COL] = pd.to_datetime(out[TIMESTAMP_COL], errors="coerce")

    out["_set_number"] = out[SET_COL].map(_safe_set_number)
    out = out.sort_values(["_set_number", TIMESTAMP_COL], kind="stable")
    out = out.drop(columns=["_set_number"]).reset_index(drop=True)

    return out


def _safe_set_number(label: object) -> float:
    """Set number for sorting. Unparseable labels sort last, then fail in scoring."""
    try:
        return float(parse_set_label(label))
    except ValueError:
        return float("inf")


def load_response_sheet(filepath: str) -> pd.DataFrame:
    """Load and canonicalize a Google Forms response export."""
    return normalize_response_sheet(pd.read_excel(filepath, sheet_name=0))


def _deduplicate_submissions(
    response_df: pd.DataFrame,
) -> tuple[pd.DataFrame, List[SupersededSubmission]]:
    """
    Keep only each student's latest submission. See docs/adr — a resubmission is
    treated as a correction, and what got dropped is reported rather than hidden.

    Rows with a blank Roll Number cannot be attributed to a student, so they are
    all kept and left for the Validation sheet to surface.
    """
    roll = response_df[ROLL_COL].astype(str).str.strip()
    identified = roll.ne("") & roll.str.lower().ne("nan")

    superseded: List[SupersededSubmission] = []
    keep_mask = pd.Series(True, index=response_df.index)

    for roll_number, group in response_df[identified].groupby(roll[identified]):
        if len(group) == 1:
            continue
        ordered = group.sort_values(TIMESTAMP_COL, kind="stable")
        kept = ordered.index[-1]
        kept_at = ordered.loc[kept, TIMESTAMP_COL]
        for idx in ordered.index[:-1]:
            keep_mask.loc[idx] = False
            superseded.append(
                SupersededSubmission(
                    roll_number=roll_number,
                    set_no=str(ordered.loc[idx, SET_COL]).strip(),
                    submitted_at=ordered.loc[idx, TIMESTAMP_COL],
                    kept_at=kept_at,
                )
            )

    return response_df[keep_mask].reset_index(drop=True), superseded


def check_all_responses(
    response_df: pd.DataFrame,
    question_papers_path: str,
    question_bank: FullQuestionBank,
    pass_threshold: float = 6.0,
) -> ScoringReport:
    """
    Validate and score all student responses.
    """
    set_to_question_nos = map_paper_to_bank_questions(question_papers_path, question_bank)

    qno_to_answer = {
        q.question_no: str(q.answer).strip().upper()
        for q in question_bank.get_all()
    }

    scored_df, superseded = _deduplicate_submissions(response_df)

    student_reports: List[StudentReport] = []

    for idx, row in scored_df.iterrows():
        set_no = str(row.get(SET_COL, "")).strip()
        if set_no not in set_to_question_nos:
            raise ValueError(
                f"Unknown or missing {SET_COL} at row {idx + 2}: '{set_no}'"
            )

        assigned_qnos = set_to_question_nos[set_no]
        assigned_set = set(assigned_qnos)
        answered = _extract_answered_questions(row)

        extra_questions = sorted(q_no for q_no in answered if q_no not in assigned_set)

        correct = 0
        wrong = 0
        unanswered = 0

        for q_no in assigned_qnos:
            answer = answered.get(q_no)
            if answer is None:
                # Compulsory forms: unanswered is treated as wrong.
                continue
            if answer == qno_to_answer[q_no]:
                correct += 1
            else:
                wrong += 1

        wrong += (len(assigned_qnos) - (correct + wrong))
        # Everything the student filled in, including questions that were not on
        # their paper. Those earn nothing, but they were still attempted.
        attempted = len(answered)
        unanswered = 0

        validation = ValidationResult(set_no=set_no, extra_questions=extra_questions)
        student_reports.append(
            StudentReport(
                student_index=idx,
                validation=validation,
                assigned=len(assigned_qnos),
                attempted=attempted,
                correct=correct,
                wrong=wrong,
                unanswered=unanswered,
                roll_number=_cell_text(row.get(ROLL_COL)),
                full_name=_cell_text(row.get(NAME_COL)),
                email=_cell_text(row.get(EMAIL_COL)),
            )
        )

    score_series = pd.Series([r.correct for r in student_reports], dtype=float)
    avg_score = round(float(score_series.mean()), 2) if len(score_series) else 0.0
    median_score = round(float(score_series.median()), 2) if len(score_series) else 0.0
    pass_count = sum(1 for r in student_reports if r.correct >= pass_threshold)
    pass_rate = round((pass_count / len(student_reports)) * 100, 2) if student_reports else 0.0

    validation_issues = [r for r in student_reports if r.validation.extra_count > 0]

    return ScoringReport(
        student_reports=student_reports,
        validation_issues=validation_issues,
        avg_score=avg_score,
        median_score=median_score,
        pass_count=pass_count,
        pass_rate=pass_rate,
        pass_threshold=pass_threshold,
        superseded=superseded,
        scored_df=scored_df,
    )


def _write_summary_charts(ws, report: ScoringReport, max_marks: int) -> None:
    """
    Add a mark-by-mark histogram with a fitted normal curve over it, beside the
    Summary metrics table (which stays in columns A-B). Chart source data goes in
    columns D-F, the chart itself at H2.

    The histogram is filled in for every mark from 0 to max, including the marks
    nobody scored, so the bars sit on an even axis and the curve reads properly.
    """
    scores = [r.correct for r in report.student_reports]
    if not scores or max_marks <= 0:
        return

    total = len(scores)
    mean = statistics.fmean(scores)
    # Population sigma: these are all the students, not a sample of them.
    sigma = statistics.pstdev(scores)

    ws["D1"] = "Marks"
    ws["E1"] = "Students"
    ws["F1"] = "Normal Curve"
    for cell in ("D1", "E1", "F1"):
        ws[cell].font = Font(bold=True)

    observed = Counter(scores)
    for offset, marks in enumerate(range(0, max_marks + 1)):
        row = 2 + offset
        ws.cell(row=row, column=4, value=marks)
        ws.cell(row=row, column=5, value=observed.get(marks, 0))
        if sigma > 0:
            density = math.exp(-((marks - mean) ** 2) / (2 * sigma**2)) / (
                sigma * math.sqrt(2 * math.pi)
            )
            # Bin width is 1 mark, so density * total is the expected head count.
            ws.cell(row=row, column=6, value=round(total * density, 2))
        else:
            # Everyone scored the same — a normal curve would be a divide by zero.
            ws.cell(row=row, column=6, value=None)

    last_row = 1 + (max_marks + 1)

    histogram = BarChart()
    histogram.type = "col"
    histogram.title = "Score Distribution vs Normal Curve"
    histogram.gapWidth = 40
    histogram.height = 10
    histogram.width = 20
    histogram.add_data(
        Reference(ws, min_col=5, min_row=1, max_row=last_row), titles_from_data=True
    )
    histogram.set_categories(Reference(ws, min_col=4, min_row=2, max_row=last_row))

    if sigma > 0:
        curve = LineChart()
        curve.add_data(
            Reference(ws, min_col=6, min_row=1, max_row=last_row), titles_from_data=True
        )
        curve.smooth = True
        histogram += curve

    # openpyxl hides both axes by default (it writes <c:delete val="1"/>), which
    # strips the tick numbers off the chart. Turn them back on explicitly.
    for axis, title in ((histogram.x_axis, "Marks obtained"),
                        (histogram.y_axis, "Number of students")):
        axis.delete = False
        axis.title = title
        axis.tickLblPos = "nextTo"
        axis.majorTickMark = "out"
        axis.numFmt = "0"

    ws.add_chart(histogram, "H2")

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 11
    ws.column_dimensions["F"].width = 14


def _color_code_answers(
    ws, first_letter: str, last_letter: str, last_row: int, mask_letter: str
) -> None:
    """
    Colour every answer cell, in the faculty's own scheme. Two independent
    signals, so a cell says both things at once:

        text        blue if the answer matches the key, red italic if it does not
        background  light green if the question was on this student's paper,
                    light yellow if they answered one that was not

    A question on their paper goes green whether they answered it or not, so the
    ones they left blank stand out as empty green cells. A question that was not
    on their paper stays unfilled unless they answered it.

    `mask_letter` is the first column of the hidden helper block, where 1 means
    "this question was on this student's paper" — it is what the two background
    rules read.

    The font rules and the fill rules set disjoint properties, so Excel applies
    one of each to the same cell without them competing.

    These are conditional formatting rules rather than fixed colours, so that
    correcting a key letter in row 2 recolours the sheet as well as recomputing
    AnsC. Formulas are written relative to the top-left cell of the range, which
    is how Excel expands them across it.
    """
    if last_row < 4:
        return

    cell_range = f"{first_letter}4:{last_letter}{last_row}"
    answer = f"{first_letter}4"
    key = f"{first_letter}$2"
    assigned = f"{mask_letter}4"

    rules = [
        # Text: right or wrong against the key, whether or not it was their question.
        (
            f'AND({answer}<>"",{answer}={key})',
            DifferentialStyle(font=Font(bold=True, color="FF0000FF")),
        ),
        (
            f'AND({answer}<>"",{answer}<>{key})',
            DifferentialStyle(font=Font(bold=True, italic=True, color="FFFF0000")),
        ),
        # Background: was this question theirs?
        (
            f"{assigned}=1",
            DifferentialStyle(fill=PatternFill(bgColor="FFC6EFCE")),
        ),
        (
            f'AND({assigned}=0,{answer}<>"")',
            DifferentialStyle(fill=PatternFill(bgColor="FFFFF2CC")),
        ),
    ]

    for formula, style in rules:
        ws.conditional_formatting.add(
            cell_range, Rule(type="expression", formula=[formula], dxf=style)
        )


def _write_faculty_report(
    writer: pd.ExcelWriter,
    response_df: pd.DataFrame,
    question_bank: FullQuestionBank,
    set_to_question_nos: Dict[str, List[int]],
) -> None:
    """
    Write the 'Faculty_Report' sheet in the layout the faculty asked for:

        row 1        question-number strip (Q01, Q02, ...)   over the answer columns
        row 2        the answer key                          over the answer columns
        row 3        headers, including Count and AnsC
        row 4..      one row per scored submission, ordered by set

    Count and AnsC are live formulas so that editing the key in row 2 recalculates
    the whole column. Their ranges cover only real question columns — see
    docs/adr/0002 before widening them.

    A hidden helper block sits to the right of the answers holding 1/0 per question
    per student: "was this question on that student's paper". AnsC multiplies by it,
    so answering a question outside your set can never earn a mark, and the purple
    conditional format reads the same block. Without it a formula comparing the row
    to the key has no way to tell an assigned question from an extra one.
    """
    ws = writer.book.create_sheet("Faculty_Report")

    question_nos = sorted(q.question_no for q in question_bank.get_all())
    qno_to_answer = {
        q.question_no: str(q.answer).strip().upper() for q in question_bank.get_all()
    }

    # Identity columns, then Count and AnsC, then one column per question.
    first_answer_col = len(IDENTITY_COLS) + len(("Count", "AnsC")) + 1
    last_answer_col = first_answer_col + len(question_nos) - 1
    first_letter = get_column_letter(first_answer_col)
    last_letter = get_column_letter(last_answer_col)

    # Hidden helper block, one blank column clear of the answers.
    mask_first_col = last_answer_col + 2
    mask_first_letter = get_column_letter(mask_first_col)
    mask_last_letter = get_column_letter(mask_first_col + len(question_nos) - 1)

    key_font = Font(bold=True, color="FF0000")
    header_font = Font(bold=True)
    ansc_font = Font(bold=True, color="0000FF")

    # Rows 1-2: the answer key, sitting directly above the columns it grades.
    for offset, q_no in enumerate(question_nos):
        col = first_answer_col + offset
        label_cell = ws.cell(row=1, column=col, value=f"Q{q_no:02d}")
        label_cell.font = key_font
        key_cell = ws.cell(row=2, column=col, value=qno_to_answer[q_no])
        key_cell.font = key_font

    # Row 3: headers.
    headers = IDENTITY_COLS + ["Count", "AnsC"] + [f"Q-{n:02d}" for n in question_nos]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = ansc_font if header == "AnsC" else header_font

    ws.cell(row=3, column=mask_first_col - 1, value="helper →").font = Font(
        bold=True, italic=True, color="FF808080"
    )

    # Rows 4+: one per submission.
    for offset, (_, row) in enumerate(response_df.iterrows()):
        excel_row = 4 + offset
        assigned_qnos = set(
            set_to_question_nos.get(_cell_text(row.get(SET_COL)), [])
        )

        ws.cell(row=excel_row, column=1, value=row.get(TIMESTAMP_COL)).number_format = (
            "yyyy-mm-dd hh:mm:ss"
        )
        ws.cell(row=excel_row, column=2, value=_cell_text(row.get(EMAIL_COL)))
        ws.cell(row=excel_row, column=3, value=_cell_text(row.get(NAME_COL)))
        ws.cell(row=excel_row, column=4, value=_cell_text(row.get(ROLL_COL)))
        ws.cell(row=excel_row, column=5, value=_cell_text(row.get(SET_COL)))

        # Everything they filled in, their own questions or not.
        ws.cell(
            row=excel_row,
            column=6,
            value=f"=COUNTA({first_letter}{excel_row}:{last_letter}{excel_row})",
        )
        # Correct answers, gated by the helper block so extras score nothing.
        ws.cell(
            row=excel_row,
            column=7,
            value=(
                f"=SUMPRODUCT(--({first_letter}$2:{last_letter}$2"
                f"={first_letter}{excel_row}:{last_letter}{excel_row}),"
                f"{mask_first_letter}{excel_row}:{mask_last_letter}{excel_row})"
            ),
        )

        for q_offset, q_no in enumerate(question_nos):
            answer = _normalize_answer(row.get(f"Q{q_no}"))
            if answer is not None:
                ws.cell(row=excel_row, column=first_answer_col + q_offset, value=answer)
            ws.cell(
                row=excel_row,
                column=mask_first_col + q_offset,
                value=1 if q_no in assigned_qnos else 0,
            )

    _color_code_answers(
        ws,
        first_letter,
        last_letter,
        last_row=3 + len(response_df),
        mask_letter=mask_first_letter,
    )

    for col in range(mask_first_col - 1, mask_first_col + len(question_nos)):
        ws.column_dimensions[get_column_letter(col)].hidden = True

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 20
    ws.freeze_panes = f"{first_letter}4"


def generate_scoring_report(
    report: ScoringReport,
    output_path: str,
    question_papers_path: Optional[str] = None,
    question_bank: Optional[FullQuestionBank] = None,
) -> str:
    """
    Write scoring report to Excel.

    Always includes:
    - Scores
    - Summary
    - Validation

    Additionally includes 'Responses_Review' (colored answer cells) and
    'Faculty_Report' (the layout the faculty asked for) when question_papers_path
    and question_bank are provided. Both are built from the submissions that were
    actually scored, so they never disagree with Scores.

    `output_path` may be a filesystem path or an in-memory buffer (e.g. io.BytesIO).
    """
    if isinstance(output_path, (str, Path)):
        Path(output_path).parent.mkdir(parents=True, exist_ok=True)

    scores_rows = []
    for r in report.student_reports:
        scores_rows.append(
            {
                "Roll Number": r.roll_number,
                "Full Name": r.full_name,
                "Email address": r.email,
                "Set": r.set_no,
                "Assigned": r.assigned,
                "Attempted": r.attempted,
                "Correct": r.correct,
                "Wrong": r.wrong,
                "Extra Answers": r.validation.extra_count,
            }
        )
    scores_df = pd.DataFrame(scores_rows)

    max_marks = scores_df["Assigned"].mode().iloc[0] if not scores_df.empty else 0

    summary_rows = [
        {"Metric": "Total Students", "Value": len(report.student_reports)},
        {"Metric": "Max Marks (per student)", "Value": max_marks},
        {"Metric": "Average Correct (marks)", "Value": report.avg_score},
        {"Metric": "Median Correct (marks)", "Value": report.median_score},
        {"Metric": f"Pass Count (≥{int(report.pass_threshold)} marks)", "Value": report.pass_count},
        {"Metric": "Pass Rate (%)", "Value": report.pass_rate},
        {"Metric": "---", "Value": "---"},
        {"Metric": "Score Distribution (Correct/Max)", "Value": None},
    ]
    for marks, count in report.grade_distribution().items():
        summary_rows.append({"Metric": marks, "Value": count})
    summary_df = pd.DataFrame(summary_rows)

    validation_rows = []
    for s in report.superseded:
        validation_rows.append(
            {
                "Roll Number": s.roll_number,
                "Set": s.set_no,
                "Issue": "Duplicate submission (dropped)",
                "Detail": f"Submitted {s.submitted_at}; scored the later one at {s.kept_at}",
            }
        )
    for r in report.validation_issues:
        extra_str = ", ".join([f"Q{q}" for q in r.validation.extra_questions])
        validation_rows.append(
            {
                "Roll Number": r.roll_number,
                "Set": r.set_no,
                "Issue": f"Answered {r.validation.extra_count} question(s) outside their set",
                "Detail": extra_str,
            }
        )

    if validation_rows:
        validation_df = pd.DataFrame(validation_rows)
    else:
        validation_df = pd.DataFrame(
            [
                {
                    "Roll Number": "",
                    "Set": "",
                    "Issue": "✅ No validation issues found",
                    "Detail": "One submission per student, and nobody answered outside their set",
                }
            ]
        )

    response_df = report.scored_df

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        scores_df.to_excel(writer, sheet_name="Scores", index=False)
        summary_df.to_excel(writer, sheet_name="Summary", index=False)
        validation_df.to_excel(writer, sheet_name="Validation", index=False)

        _write_summary_charts(writer.book["Summary"], report, max_marks)

        if (
            response_df is not None
            and question_papers_path is not None
            and question_bank is not None
        ):
            set_to_question_nos = map_paper_to_bank_questions(question_papers_path, question_bank)

            _write_faculty_report(
                writer, response_df, question_bank, set_to_question_nos
            )

            review_df = response_df.copy()
            review_df.to_excel(writer, sheet_name="Responses_Review", index=False)

            ws = writer.book["Responses_Review"]
            green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            # Out-of-set answers: attempted, but worth nothing either way.
            purple_fill = PatternFill(start_color="E4D7F5", end_color="E4D7F5", fill_type="solid")
            purple_font = Font(bold=True, color="FF7030A0")

            qno_to_answer = {
                q.question_no: str(q.answer).strip().upper()
                for q in question_bank.get_all()
            }

            col_to_idx = {str(col): idx + 1 for idx, col in enumerate(review_df.columns)}

            for row_idx, row in review_df.iterrows():
                set_no = str(row.get(SET_COL, "")).strip()
                assigned_qnos = set(set_to_question_nos.get(set_no, []))
                excel_row = row_idx + 2  # Header is row 1

                for col in review_df.columns:
                    match = QUESTION_COL_RE.match(str(col))
                    if not match:
                        continue

                    q_no = int(match.group(1))
                    answer = _normalize_answer(row[col])
                    if answer is None:
                        continue

                    excel_col = col_to_idx[str(col)]
                    cell = ws.cell(row=excel_row, column=excel_col)

                    if q_no not in assigned_qnos:
                        # Not on their paper — never marked, right or wrong.
                        cell.fill = purple_fill
                        cell.font = purple_font
                    elif answer == qno_to_answer.get(q_no):
                        cell.fill = green_fill
                    else:
                        cell.fill = red_fill

    return output_path
