"""
Quiz Generator - Streamlit UI

A web interface for generating randomized quiz papers from a question bank.
Upload an Excel file, configure settings, and download formatted question papers.
"""

import streamlit as st
import pandas as pd
import numpy as np
import io
import os
import secrets
import tempfile
from typing import Dict, List
from collections import defaultdict

from allocator import QuizStructure, allocate_quizzes, shuffle_all_quizzes
from excel_handler import load_question_bank, FullQuestionBank
from response_generator import generate_responses
from answer_checker import (
    load_response_sheet,
    check_all_responses,
    generate_scoring_report,
)
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill


# Page configuration
st.set_page_config(
    page_title="Quiz Generator",
    page_icon="📝",
    layout="wide"
)


def qid_to_number(question_id: str, question_bank: FullQuestionBank) -> int:
    """Convert internal question_id (e.g., H1, M5) to original question_no."""
    q = question_bank.get_by_id(question_id)
    return q.question_no if q else 0


def create_formatted_excel(
    allocation_matrix: list,
    shuffled_matrix: list,
    usage_counts: dict,
    question_bank: FullQuestionBank,
    include_answer_key: bool = True
) -> bytes:
    """
    Create a formatted Excel file with:
    - Question papers (one sheet per student)
    - Answer Key
    - Allocation Table (original order, numeric IDs)
    - Shuffled Table (shuffled order, numeric IDs)
    - Evaluation Table (min/max/delta stats)
    """
    wb = Workbook()

    # ── Shared styles ─────────────────────────────────────────────────────
    header_font = Font(bold=True, size=14)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    green_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    orange_fill = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    wrap_align = Alignment(wrap_text=True, vertical='top')
    center_align = Alignment(horizontal='center', vertical='center')

    # Remove default sheet
    wb.remove(wb.active)

    # ══════════════════════════════════════════════════════════════════════
    # Question Paper Sheets (one per student)
    # ══════════════════════════════════════════════════════════════════════
    for student_idx, quiz in enumerate(shuffled_matrix):
        ws = wb.create_sheet(title=f"Set_{student_idx + 1}")

        # Title
        ws.merge_cells('A1:F1')
        ws['A1'] = f"Question Paper - Set {student_idx + 1}"
        ws['A1'].font = header_font
        ws['A1'].alignment = center_align

        # Headers
        for col, header in enumerate(['Q.No', 'Question', 'Option A', 'Option B', 'Option C', 'Option D'], 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = center_align

        # Questions
        for q_idx, question_id in enumerate(quiz):
            q = question_bank.get_by_id(question_id)
            row = q_idx + 4
            ws.cell(row=row, column=1, value=q_idx + 1).border = thin_border
            ws.cell(row=row, column=1).alignment = center_align
            ws.cell(row=row, column=2, value=q.question_text).border = thin_border
            ws.cell(row=row, column=2).alignment = wrap_align
            ws.cell(row=row, column=3, value=q.option_a).border = thin_border
            ws.cell(row=row, column=4, value=q.option_b).border = thin_border
            ws.cell(row=row, column=5, value=q.option_c).border = thin_border
            ws.cell(row=row, column=6, value=q.option_d).border = thin_border

        # Column widths & row heights
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 50
        for ch in 'CDEF':
            ws.column_dimensions[ch].width = 20
        for r in range(4, len(quiz) + 4):
            ws.row_dimensions[r].height = 30

    # ══════════════════════════════════════════════════════════════════════
    # Answer Key Sheet
    # ══════════════════════════════════════════════════════════════════════
    if include_answer_key:
        ws = wb.create_sheet(title="Answer_Key")
        ws['A1'] = "ANSWER KEY (For Teachers Only)"
        ws['A1'].font = Font(bold=True, size=16, color="FF0000")
        num_q = len(shuffled_matrix[0])
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_q + 1)

        headers = ['Set'] + [f'Q{i+1}' for i in range(num_q)]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=h)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.border = thin_border

        for student_idx, quiz in enumerate(shuffled_matrix):
            row = student_idx + 4
            ws.cell(row=row, column=1, value=f"Set_{student_idx + 1}").border = thin_border
            for q_idx, qid in enumerate(quiz):
                q = question_bank.get_by_id(qid)
                ws.cell(row=row, column=q_idx + 2, value=q.answer).border = thin_border

    # ══════════════════════════════════════════════════════════════════════
    # Allocation Table Sheet (original order, numeric question numbers)
    # ══════════════════════════════════════════════════════════════════════
    ws = wb.create_sheet(title="Allocation_Table")
    ws['A1'] = "Allocation Table (Original Order by Difficulty)"
    ws['A1'].font = Font(bold=True, size=14)
    num_students = len(allocation_matrix)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_students + 1)

    # Headers
    cell = ws.cell(row=3, column=1, value="Position")
    cell.font = header_font_white
    cell.fill = header_fill
    cell.border = thin_border
    for s_idx in range(num_students):
        cell = ws.cell(row=3, column=s_idx + 2, value=f"S{s_idx + 1}")
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border

    # Data (using original question_no, not H/M/E IDs)
    num_positions = len(allocation_matrix[0])
    for pos in range(num_positions):
        cell = ws.cell(row=pos + 4, column=1, value=f"Q{pos + 1}")
        cell.border = thin_border
        cell.font = Font(bold=True)
        for s_idx in range(num_students):
            qid = allocation_matrix[s_idx][pos]
            cell = ws.cell(row=pos + 4, column=s_idx + 2, value=qid_to_number(qid, question_bank))
            cell.border = thin_border
            cell.alignment = center_align

    ws.column_dimensions['A'].width = 10

    # ══════════════════════════════════════════════════════════════════════
    # Shuffled Table Sheet (shuffled order, numeric question numbers)
    # ══════════════════════════════════════════════════════════════════════
    ws = wb.create_sheet(title="Shuffled_Table")
    ws['A1'] = "Shuffled Table (Randomized Order per Student)"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_students + 1)

    # Headers
    cell = ws.cell(row=3, column=1, value="Position")
    cell.font = header_font_white
    cell.fill = green_fill
    cell.border = thin_border
    for s_idx in range(num_students):
        cell = ws.cell(row=3, column=s_idx + 2, value=f"S{s_idx + 1}")
        cell.font = header_font_white
        cell.fill = green_fill
        cell.border = thin_border

    # Data
    for pos in range(num_positions):
        cell = ws.cell(row=pos + 4, column=1, value=f"Q{pos + 1}")
        cell.border = thin_border
        cell.font = Font(bold=True)
        for s_idx in range(num_students):
            qid = shuffled_matrix[s_idx][pos]
            cell = ws.cell(row=pos + 4, column=s_idx + 2, value=qid_to_number(qid, question_bank))
            cell.border = thin_border
            cell.alignment = center_align

    ws.column_dimensions['A'].width = 10

    # ══════════════════════════════════════════════════════════════════════
    # Evaluation Table Sheet
    # ══════════════════════════════════════════════════════════════════════
    ws = wb.create_sheet(title="Evaluation")
    ws['A1'] = "Evaluation Summary"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:E1')

    # ── Question Usage Table ──
    ws['A3'] = "Question Usage"
    ws['A3'].font = Font(bold=True, size=12)

    for col, h in enumerate(['Question No', 'Internal ID', 'Difficulty', 'Usage Count'], 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border

    row = 5
    for q in question_bank.get_all():
        count = usage_counts.get(q.question_id, 0)
        ws.cell(row=row, column=1, value=q.question_no).border = thin_border
        ws.cell(row=row, column=1).alignment = center_align
        ws.cell(row=row, column=2, value=q.question_id).border = thin_border
        ws.cell(row=row, column=3, value=q.difficulty.capitalize()).border = thin_border
        ws.cell(row=row, column=4, value=count).border = thin_border
        ws.cell(row=row, column=4).alignment = center_align
        row += 1

    # ── Min/Max/Delta by Difficulty ──
    row += 1
    ws.cell(row=row, column=1, value="Min / Max / Delta by Difficulty").font = Font(bold=True, size=12)
    row += 1

    for col, h in enumerate(['Difficulty', 'Min', 'Max', 'Delta', 'Variance'], 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = header_font_white
        cell.fill = orange_fill
        cell.border = thin_border
    row += 1

    by_diff = defaultdict(list)
    for q in question_bank.get_all():
        by_diff[q.difficulty].append(usage_counts.get(q.question_id, 0))

    all_counts = list(usage_counts.values())

    # Track min/max per difficulty for overall calculation
    diff_stats = []
    
    for diff in ['hard', 'medium', 'easy']:
        counts = by_diff.get(diff, [])
        if counts:
            mn, mx = min(counts), max(counts)
            var = round(float(np.var(counts)), 4)
        else:
            mn = mx = 0
            var = 0.0
        
        diff_stats.append((mn, mx))
        
        ws.cell(row=row, column=1, value=diff.capitalize()).border = thin_border
        ws.cell(row=row, column=2, value=mn).border = thin_border
        ws.cell(row=row, column=3, value=mx).border = thin_border
        ws.cell(row=row, column=4, value=mx - mn).border = thin_border
        ws.cell(row=row, column=5, value=var).border = thin_border
        row += 1

    # Overall row: sum of min/max from each difficulty
    overall_min = sum(mn for mn, mx in diff_stats)
    overall_max = sum(mx for mn, mx in diff_stats)
    overall_delta = overall_max - overall_min
    
    ws.cell(row=row, column=1, value="OVERALL").border = thin_border
    ws.cell(row=row, column=1).font = Font(bold=True)
    ws.cell(row=row, column=2, value=overall_min).border = thin_border
    ws.cell(row=row, column=3, value=overall_max).border = thin_border
    ws.cell(row=row, column=4, value=overall_delta).border = thin_border
    ws.cell(row=row, column=5, value="-").border = thin_border

    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 12

    # ══════════════════════════════════════════════════════════════════════
    # Question Bank Sheet (embedded for Part 2 answer checking)
    # ══════════════════════════════════════════════════════════════════════
    ws = wb.create_sheet(title="Question_Bank")
    ws['A1'] = "Question Bank (Embedded for Answer Checking)"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:H1')

    qb_headers = ['question_no', 'question', 'option_a', 'option_b',
                   'option_c', 'option_d', 'answer', 'difficulty']
    qb_fill = PatternFill(start_color="8DB4E2", end_color="8DB4E2", fill_type="solid")
    for col, h in enumerate(qb_headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = header_font_white
        cell.fill = qb_fill
        cell.border = thin_border

    for q_idx, q in enumerate(question_bank.get_all()):
        row = q_idx + 4
        ws.cell(row=row, column=1, value=q.question_no).border = thin_border
        ws.cell(row=row, column=1).alignment = center_align
        ws.cell(row=row, column=2, value=q.question_text).border = thin_border
        ws.cell(row=row, column=2).alignment = wrap_align
        ws.cell(row=row, column=3, value=q.option_a).border = thin_border
        ws.cell(row=row, column=4, value=q.option_b).border = thin_border
        ws.cell(row=row, column=5, value=q.option_c).border = thin_border
        ws.cell(row=row, column=6, value=q.option_d).border = thin_border
        ws.cell(row=row, column=7, value=q.answer).border = thin_border
        ws.cell(row=row, column=7).alignment = center_align
        ws.cell(row=row, column=8, value=q.difficulty.capitalize()).border = thin_border

    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 50
    for ch in 'CDEF':
        ws.column_dimensions[ch].width = 20
    ws.column_dimensions['G'].width = 10
    ws.column_dimensions['H'].width = 12

    # ── Save ──────────────────────────────────────────────────────────────
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def _save_uploaded_temp(uploaded_file, prefix: str) -> str:
    """Persist uploaded Streamlit file to a temporary .xlsx path."""
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx", prefix=prefix) as tmp:
        tmp.write(uploaded_file.getvalue())
        return tmp.name


def _make_excel_bytes_from_dataframe(df: pd.DataFrame, sheet_name: str) -> bytes:
    """Serialize DataFrame into an in-memory Excel file."""
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)
    output.seek(0)
    return output.getvalue()


def _render_generation_tab():
    """Part 1 UI: generate question papers."""
    st.markdown("Upload a question bank and generate randomized question papers for all students.")
    st.divider()

    st.header("1️⃣ Upload Question Bank")
    uploaded_file = st.file_uploader(
        "Upload Excel file (.xlsx)",
        type=["xlsx"],
        help="Excel with columns: question_no, question, option_a, option_b, option_c, option_d, answer, difficulty",
        key="part1_question_bank",
    )

    question_bank = None
    counts = {"hard": 0, "medium": 0, "easy": 0}

    if uploaded_file:
        temp_path = None
        try:
            temp_path = _save_uploaded_temp(uploaded_file, "part1_qb_")
            question_bank = load_question_bank(temp_path)
            counts = question_bank.count_by_difficulty()
            total = sum(counts.values())

            st.success(f"✅ Loaded {total} questions successfully!")
            col1, col2, col3, col4 = st.columns(4)
            col1.metric("Hard", counts.get("hard", 0))
            col2.metric("Medium", counts.get("medium", 0))
            col3.metric("Easy", counts.get("easy", 0))
            col4.metric("Total", total)

            st.session_state["part1_question_bank_bytes"] = uploaded_file.getvalue()
        except Exception as e:
            st.error(f"❌ Error loading file: {str(e)}")
            question_bank = None
        finally:
            if temp_path and os.path.exists(temp_path):
                os.remove(temp_path)

    st.divider()

    st.header("2️⃣ Configuration")
    col1, col2 = st.columns(2)
    with col1:
        num_students = st.number_input(
            "Number of Students",
            min_value=1,
            max_value=500,
            value=50,
            help="How many question papers to generate",
            key="part1_num_students",
        )
    with col2:
        total_questions = st.number_input(
            "Questions per Quiz",
            min_value=1,
            max_value=100,
            value=15,
            help="Total questions in each student's quiz",
            key="part1_total_questions",
        )

    st.divider()
    st.header("3️⃣ Difficulty Distribution")
    mode = st.radio(
        "Select mode:",
        ["Absolute (exact counts)", "Percentage (auto-calculate)"],
        horizontal=True,
        key="part1_mode",
    )

    hard_count = medium_count = easy_count = 0
    if mode == "Absolute (exact counts)":
        col1, col2, col3 = st.columns(3)
        with col1:
            hard_count = st.number_input(
                "Hard Questions",
                min_value=0,
                max_value=total_questions,
                value=min(4, counts.get("hard", 4)),
                key="part1_hard_abs",
            )
        with col2:
            medium_count = st.number_input(
                "Medium Questions",
                min_value=0,
                max_value=total_questions,
                value=min(6, counts.get("medium", 6)),
                key="part1_medium_abs",
            )
        with col3:
            easy_count = st.number_input(
                "Easy Questions",
                min_value=0,
                max_value=total_questions,
                value=min(5, counts.get("easy", 5)),
                key="part1_easy_abs",
            )
    else:
        col1, col2, col3 = st.columns(3)
        with col1:
            hard_pct = st.slider("Hard %", 0, 100, 27, key="part1_hard_pct")
        with col2:
            medium_pct = st.slider("Medium %", 0, 100, 40, key="part1_medium_pct")
        with col3:
            easy_pct = st.slider("Easy %", 0, 100, 33, key="part1_easy_pct")

        total_pct = hard_pct + medium_pct + easy_pct
        if total_pct != 100:
            st.warning(f"⚠️ Percentages sum to {total_pct}%, should be 100%")

        hard_count = round(total_questions * hard_pct / 100)
        medium_count = round(total_questions * medium_pct / 100)
        easy_count = total_questions - hard_count - medium_count
        st.info(
            f"📊 Calculated: {hard_count} Hard + {medium_count} Medium + {easy_count} Easy = "
            f"{hard_count + medium_count + easy_count} questions"
        )

    total_selected = hard_count + medium_count + easy_count
    if total_selected != total_questions:
        st.error(f"❌ Selected {total_selected} questions, but quiz requires {total_questions}")

    validation_errors = []
    if question_bank:
        if hard_count > counts.get("hard", 0):
            validation_errors.append(f"Need {hard_count} hard questions, only {counts.get('hard', 0)} available")
        if medium_count > counts.get("medium", 0):
            validation_errors.append(
                f"Need {medium_count} medium questions, only {counts.get('medium', 0)} available"
            )
        if easy_count > counts.get("easy", 0):
            validation_errors.append(f"Need {easy_count} easy questions, only {counts.get('easy', 0)} available")

    for error in validation_errors:
        st.error(f"❌ {error}")

    st.divider()
    st.header("4️⃣ Randomization")
    use_fixed_seed = st.checkbox(
        "Use fixed seed (reproducible output)",
        value=False,
        help="Enable this if you want the exact same allocation for identical inputs.",
        key="part1_use_fixed_seed",
    )
    fixed_seed = None
    if use_fixed_seed:
        fixed_seed = st.number_input(
            "Seed value",
            min_value=0,
            max_value=2_147_483_647,
            value=42,
            step=1,
            key="part1_seed_value",
        )
        st.caption("Same input + same seed -> same allocation.")
    else:
        st.caption("Each generation uses a fresh random seed.")

    st.divider()
    st.header("5️⃣ Generate Question Papers")

    can_generate = question_bank is not None and total_selected == total_questions and len(validation_errors) == 0

    if st.button("🚀 Generate Question Papers", disabled=not can_generate, type="primary", key="part1_generate"):
        with st.spinner("Generating question papers..."):
            try:
                quiz_structure = QuizStructure(
                    hard_count=hard_count,
                    medium_count=medium_count,
                    easy_count=easy_count,
                )
                q_ids_by_diff = {
                    "hard": question_bank.get_question_ids_by_difficulty("hard"),
                    "medium": question_bank.get_question_ids_by_difficulty("medium"),
                    "easy": question_bank.get_question_ids_by_difficulty("easy"),
                }

                run_seed = int(fixed_seed) if use_fixed_seed else secrets.randbelow(2_147_483_647)
                allocation_matrix, usage_counts = allocate_quizzes(
                    q_ids_by_diff,
                    num_students=num_students,
                    quiz_structure=quiz_structure,
                    seed=run_seed,
                )
                shuffled_matrix = shuffle_all_quizzes(allocation_matrix, base_seed=run_seed)

                excel_bytes = create_formatted_excel(
                    allocation_matrix=allocation_matrix,
                    shuffled_matrix=shuffled_matrix,
                    usage_counts=usage_counts,
                    question_bank=question_bank,
                )

                st.session_state["part1_question_papers_bytes"] = excel_bytes
                st.success(f"✅ Generated {num_students} question papers!")
                st.download_button(
                    label="📥 Download Question Papers (Excel)",
                    data=excel_bytes,
                    file_name="question_papers.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    type="primary",
                    key="part1_download_papers",
                )

                col1, col2, col3 = st.columns(3)
                col1.metric("Students", num_students)
                col2.metric("Questions/Quiz", total_questions)
                col3.metric("Total Sheets", num_students + 5)  # sets + answer key + alloc + shuffled + eval + qbank

                if use_fixed_seed:
                    st.caption(f"Seed used: {run_seed} (fixed)")
                else:
                    st.caption(f"Seed used: {run_seed} (auto-generated for this run)")

                st.caption(
                    "Sheets: Set_1 … Set_N, Answer_Key, Allocation_Table, Shuffled_Table, Evaluation, Question_Bank"
                )
            except Exception as e:
                st.error(f"❌ Error: {str(e)}")


def _render_answer_checking_tab():
    """Part 2 UI: generate responses and score submissions."""
    st.markdown("Generate dummy responses and validate/score answer sheets.")
    st.divider()

    st.header("A️⃣ Generate Dummy Responses")
    use_part1_assets = st.checkbox(
        "Use Question Bank + Question Papers from Part 1 tab (if generated in this session)",
        value=bool(
            st.session_state.get("part1_question_bank_bytes") and st.session_state.get("part1_question_papers_bytes")
        ),
        key="part2_use_part1_assets",
    )

    gen_qb_upload = None
    gen_qp_upload = None
    if not use_part1_assets:
        gen_qb_upload = st.file_uploader("Upload Question Bank (.xlsx)", type=["xlsx"], key="part2_gen_qb_upload")
        gen_qp_upload = st.file_uploader("Upload Question Papers (.xlsx)", type=["xlsx"], key="part2_gen_qp_upload")
    else:
        if not st.session_state.get("part1_question_bank_bytes") or not st.session_state.get("part1_question_papers_bytes"):
            st.warning("Part 1 files are not available in session. Upload files manually.")
            use_part1_assets = False
            gen_qb_upload = st.file_uploader("Upload Question Bank (.xlsx)", type=["xlsx"], key="part2_gen_qb_fallback")
            gen_qp_upload = st.file_uploader(
                "Upload Question Papers (.xlsx)", type=["xlsx"], key="part2_gen_qp_fallback"
            )

    col1, col2, col3 = st.columns(3)
    with col1:
        gen_students = st.number_input("Students", min_value=1, max_value=500, value=70, key="part2_gen_students")
    with col2:
        correct_rate = st.slider("Correct %", 0, 100, 70, key="part2_correct_rate")
    with col3:
        wrong_rate = st.slider("Wrong %", 0, 100, 20, key="part2_wrong_rate")

    blank_rate = 100 - correct_rate - wrong_rate
    if blank_rate < 0:
        st.error("❌ Correct% + Wrong% cannot exceed 100.")
    else:
        st.caption(f"Blank % (auto): {blank_rate}")

    use_fixed_gen_seed = st.checkbox(
        "Use fixed seed for dummy responses",
        value=False,
        key="part2_use_fixed_gen_seed",
    )
    gen_seed = None
    if use_fixed_gen_seed:
        gen_seed = st.number_input(
            "Generator seed",
            min_value=0,
            max_value=2_147_483_647,
            value=42,
            step=1,
            key="part2_gen_seed",
        )

    if st.button("🧪 Generate Dummy Responses", type="primary", key="part2_generate_responses"):
        if blank_rate < 0:
            st.error("Fix rates before generating responses.")
        else:
            temp_files = []
            try:
                if use_part1_assets:
                    qb_bytes = st.session_state["part1_question_bank_bytes"]
                    qp_bytes = st.session_state["part1_question_papers_bytes"]
                else:
                    if not gen_qb_upload or not gen_qp_upload:
                        st.error("Upload both Question Bank and Question Papers.")
                        return
                    qb_bytes = gen_qb_upload.getvalue()
                    qp_bytes = gen_qp_upload.getvalue()

                with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx", prefix="part2_qb_") as qb_tmp:
                    qb_tmp.write(qb_bytes)
                    qb_path = qb_tmp.name
                with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx", prefix="part2_qp_") as qp_tmp:
                    qp_tmp.write(qp_bytes)
                    qp_path = qp_tmp.name
                temp_files.extend([qb_path, qp_path])

                question_bank = load_question_bank(qb_path)
                response_df = generate_responses(
                    question_papers_path=qp_path,
                    question_bank=question_bank,
                    num_students=int(gen_students),
                    correct_rate=float(correct_rate) / 100.0,
                    wrong_rate=float(wrong_rate) / 100.0,
                    blank_rate=float(blank_rate) / 100.0,
                    seed=int(gen_seed) if use_fixed_gen_seed else None,
                )

                response_bytes = _make_excel_bytes_from_dataframe(response_df, "Responses")
                st.session_state["part2_generated_responses_bytes"] = response_bytes

                st.success(f"✅ Generated dummy responses for {len(response_df)} students.")
                st.download_button(
                    "📥 Download student_responses.xlsx",
                    data=response_bytes,
                    file_name="student_responses.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="part2_download_responses",
                )
                st.caption(f"Shape: {response_df.shape[0]} rows × {response_df.shape[1]} columns")
            except Exception as e:
                st.error(f"❌ Error: {str(e)}")
            finally:
                for path in temp_files:
                    if os.path.exists(path):
                        os.remove(path)

    st.divider()
    st.header("B️⃣ Check & Score Responses")

    use_generated_responses = st.checkbox(
        "Use generated responses from section A (if available)",
        value=bool(st.session_state.get("part2_generated_responses_bytes")),
        key="part2_use_generated_responses",
    )

    chk_qb_upload = st.file_uploader("Upload Question Bank (.xlsx)", type=["xlsx"], key="part2_chk_qb_upload")
    chk_qp_upload = st.file_uploader("Upload Question Papers (.xlsx)", type=["xlsx"], key="part2_chk_qp_upload")
    chk_resp_upload = None
    if not use_generated_responses:
        chk_resp_upload = st.file_uploader("Upload Student Responses (.xlsx)", type=["xlsx"], key="part2_chk_resp_upload")

    pass_threshold = st.number_input(
        "Pass Threshold (%)",
        min_value=0.0,
        max_value=100.0,
        value=40.0,
        step=1.0,
        key="part2_pass_threshold",
    )

    if st.button("✅ Check & Score", type="primary", key="part2_check_score"):
        if not chk_qb_upload or not chk_qp_upload:
            st.error("Upload Question Bank and Question Papers for checking.")
        elif (not use_generated_responses) and (not chk_resp_upload):
            st.error("Upload Student Responses or enable generated responses.")
        else:
            temp_files = []
            report_temp_path = None
            try:
                qb_bytes = chk_qb_upload.getvalue()
                qp_bytes = chk_qp_upload.getvalue()
                if use_generated_responses:
                    if not st.session_state.get("part2_generated_responses_bytes"):
                        st.error("No generated responses in session. Upload response file instead.")
                        return
                    resp_bytes = st.session_state["part2_generated_responses_bytes"]
                else:
                    resp_bytes = chk_resp_upload.getvalue()

                with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx", prefix="part2_chk_qb_") as qb_tmp:
                    qb_tmp.write(qb_bytes)
                    qb_path = qb_tmp.name
                with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx", prefix="part2_chk_qp_") as qp_tmp:
                    qp_tmp.write(qp_bytes)
                    qp_path = qp_tmp.name
                with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx", prefix="part2_chk_resp_") as resp_tmp:
                    resp_tmp.write(resp_bytes)
                    resp_path = resp_tmp.name
                temp_files.extend([qb_path, qp_path, resp_path])

                question_bank = load_question_bank(qb_path)
                response_df = load_response_sheet(resp_path)
                report = check_all_responses(
                    response_df=response_df,
                    question_papers_path=qp_path,
                    question_bank=question_bank,
                    pass_threshold=float(pass_threshold),
                )

                with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx", prefix="part2_report_") as report_tmp:
                    report_temp_path = report_tmp.name
                generate_scoring_report(report, report_temp_path)
                with open(report_temp_path, "rb") as f:
                    report_bytes = f.read()

                st.success("✅ Scoring completed.")
                col1, col2, col3 = st.columns(3)
                col1.metric("Average Score (%)", f"{report.avg_score:.2f}")
                col2.metric("Median Score (%)", f"{report.median_score:.2f}")
                col3.metric("Pass Rate (%)", f"{report.pass_rate:.2f}")

                if report.validation_issues:
                    st.warning(f"⚠️ Validation issues found for {len(report.validation_issues)} students.")
                else:
                    st.success("No validation issues found.")

                st.download_button(
                    "📥 Download scoring_report.xlsx",
                    data=report_bytes,
                    file_name="scoring_report.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="part2_download_report",
                )
            except Exception as e:
                st.error(f"❌ Error: {str(e)}")
            finally:
                for path in temp_files:
                    if os.path.exists(path):
                        os.remove(path)
                if report_temp_path and os.path.exists(report_temp_path):
                    os.remove(report_temp_path)


def main():
    st.title("📝 Quiz Generator")
    part1_tab, part2_tab = st.tabs(["Part 1: Generate Papers", "Part 2: Answer Checking"])

    with part1_tab:
        _render_generation_tab()

    with part2_tab:
        _render_answer_checking_tab()


if __name__ == "__main__":
    main()
