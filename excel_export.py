"""
Excel Export Helpers

UI-agnostic functions shared by the Streamlit (app.py) and Reflex (quiz_web) UIs:
- Building the styled multi-sheet question-papers workbook.
- Serializing a DataFrame to in-memory Excel bytes.
- Loading the embedded Question_Bank sheet back out of a question papers file.
"""

import io
import os
import tempfile
from collections import defaultdict

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

from excel_handler import load_question_bank, set_label, FullQuestionBank


def qid_to_number(question_id: str, question_bank: FullQuestionBank) -> int:
    """Convert internal question_id (e.g., H1, M5) to original question_no."""
    q = question_bank.get_by_id(question_id)
    return q.question_no if q else 0


def question_code(question_no: int) -> str:
    """Printed form of a Question Number, e.g. 27 -> 'Q- 27'."""
    return f"Q- {question_no:02d}"


def create_formatted_excel(
    allocation_matrix: list,
    shuffled_matrix: list,
    usage_counts: dict,
    question_bank: FullQuestionBank,
    include_answer_key: bool = True
) -> bytes:
    """
    Create a formatted Excel file with:
    - Question papers (one sheet per student)
    - Answer Key
    - Allocation Table (original order, numeric IDs)
    - Shuffled Table (shuffled order, numeric IDs)
    - Evaluation Table (min/max/delta stats)
    """
    wb = Workbook()

    # ── Shared styles ─────────────────────────────────────────────────────
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    green_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    orange_fill = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    wrap_align = Alignment(wrap_text=True, vertical='top')
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')
    bold_font = Font(bold=True)
    qset_font = Font(bold=True, color="FF0000")
    qset_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    qcd_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")

    # Remove default sheet
    wb.remove(wb.active)

    # ══════════════════════════════════════════════════════════════════════
    # Question Paper Sheets (one per student)
    # ══════════════════════════════════════════════════════════════════════
    for student_idx, quiz in enumerate(shuffled_matrix):
        label = set_label(student_idx + 1)
        ws = wb.create_sheet(title=label)

        # Row 1: set label, then blank Name / Roll No fields to fill in by hand
        for col in (1, 2, 3):
            cell = ws.cell(row=1, column=col, value="QSet:" if col == 1 else label)
            cell.font = qset_font
            cell.fill = qset_fill
            cell.border = thin_border
            cell.alignment = center_align

        ws.merge_cells(start_row=1, start_column=4, end_row=1, end_column=6)
        ws.merge_cells(start_row=1, start_column=7, end_row=1, end_column=8)
        ws.cell(row=1, column=4, value="Name:").font = bold_font
        ws.cell(row=1, column=7, value="Roll No:").font = bold_font
        for col in range(4, 9):
            ws.cell(row=1, column=col).border = thin_border
            ws.cell(row=1, column=col).alignment = left_align

        # Row 2: headers. Two QCd columns — the bare Question Number and its
        # printed 'Q- 27' form, which is what students copy into the form.
        for col, header in enumerate(['Sr', 'QCd', 'QCd', 'Question', 'A', 'B', 'C', 'D'], 1):
            cell = ws.cell(row=2, column=col, value=header)
            cell.font = bold_font
            cell.border = thin_border
            cell.alignment = center_align
        ws.cell(row=2, column=2).fill = qcd_fill

        # Questions
        for q_idx, question_id in enumerate(quiz):
            q = question_bank.get_by_id(question_id)
            row = q_idx + 3
            ws.cell(row=row, column=1, value=q_idx + 1)
            ws.cell(row=row, column=2, value=q.question_no).fill = qcd_fill
            ws.cell(row=row, column=3, value=question_code(q.question_no))
            ws.cell(row=row, column=4, value=q.question_text)
            options = (q.option_a, q.option_b, q.option_c, q.option_d)
            for col, (letter, text) in enumerate(zip('ABCD', options), 5):
                ws.cell(row=row, column=col, value=f"{letter}] {text}")

            for col in range(1, 9):
                cell = ws.cell(row=row, column=col)
                cell.border = thin_border
                cell.alignment = center_align if col <= 3 else wrap_align

        # Column widths & row heights
        for ch, width in zip('ABCDEFGH', (6, 8, 10, 55, 28, 28, 28, 28)):
            ws.column_dimensions[ch].width = width
        for r in range(3, len(quiz) + 3):
            ws.row_dimensions[r].height = 45

    # ══════════════════════════════════════════════════════════════════════
    # Answer Key Sheet
    # ══════════════════════════════════════════════════════════════════════
    if include_answer_key:
        ws = wb.create_sheet(title="Answer_Key")
        ws['A1'] = "ANSWER KEY (For Teachers Only)"
        ws['A1'].font = Font(bold=True, size=16, color="FF0000")
        num_q = len(shuffled_matrix[0])
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_q + 1)

        headers = ['Set'] + [f'Q{i+1}' for i in range(num_q)]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=h)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.border = thin_border

        for student_idx, quiz in enumerate(shuffled_matrix):
            row = student_idx + 4
            ws.cell(row=row, column=1, value=set_label(student_idx + 1)).border = thin_border
            for q_idx, qid in enumerate(quiz):
                q = question_bank.get_by_id(qid)
                ws.cell(row=row, column=q_idx + 2, value=q.answer).border = thin_border

    # ══════════════════════════════════════════════════════════════════════
    # Allocation Table Sheet (original order, numeric question numbers)
    # ══════════════════════════════════════════════════════════════════════
    ws = wb.create_sheet(title="Allocation_Table")
    ws['A1'] = "Allocation Table (Original Order by Difficulty)"
    ws['A1'].font = Font(bold=True, size=14)
    num_students = len(allocation_matrix)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_students + 1)

    # Headers
    cell = ws.cell(row=3, column=1, value="Position")
    cell.font = header_font_white
    cell.fill = header_fill
    cell.border = thin_border
    for s_idx in range(num_students):
        cell = ws.cell(row=3, column=s_idx + 2, value=set_label(s_idx + 1))
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border

    # Data (using original question_no, not H/M/E IDs)
    num_positions = len(allocation_matrix[0])
    for pos in range(num_positions):
        cell = ws.cell(row=pos + 4, column=1, value=f"Q{pos + 1}")
        cell.border = thin_border
        cell.font = Font(bold=True)
        for s_idx in range(num_students):
            qid = allocation_matrix[s_idx][pos]
            cell = ws.cell(row=pos + 4, column=s_idx + 2, value=qid_to_number(qid, question_bank))
            cell.border = thin_border
            cell.alignment = center_align

    ws.column_dimensions['A'].width = 10

    # ══════════════════════════════════════════════════════════════════════
    # Shuffled Table Sheet (shuffled order, numeric question numbers)
    # ══════════════════════════════════════════════════════════════════════
    ws = wb.create_sheet(title="Shuffled_Table")
    ws['A1'] = "Shuffled Table (Randomized Order per Student)"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_students + 1)

    # Headers
    cell = ws.cell(row=3, column=1, value="Position")
    cell.font = header_font_white
    cell.fill = green_fill
    cell.border = thin_border
    for s_idx in range(num_students):
        cell = ws.cell(row=3, column=s_idx + 2, value=set_label(s_idx + 1))
        cell.font = header_font_white
        cell.fill = green_fill
        cell.border = thin_border

    # Data
    for pos in range(num_positions):
        cell = ws.cell(row=pos + 4, column=1, value=f"Q{pos + 1}")
        cell.border = thin_border
        cell.font = Font(bold=True)
        for s_idx in range(num_students):
            qid = shuffled_matrix[s_idx][pos]
            cell = ws.cell(row=pos + 4, column=s_idx + 2, value=qid_to_number(qid, question_bank))
            cell.border = thin_border
            cell.alignment = center_align

    ws.column_dimensions['A'].width = 10

    # ══════════════════════════════════════════════════════════════════════
    # Evaluation Table Sheet
    # ══════════════════════════════════════════════════════════════════════
    ws = wb.create_sheet(title="Evaluation")
    ws['A1'] = "Evaluation Summary"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:E1')

    # ── Question Usage Table ──
    ws['A3'] = "Question Usage"
    ws['A3'].font = Font(bold=True, size=12)

    for col, h in enumerate(['Question No', 'Internal ID', 'Difficulty', 'Usage Count'], 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border

    row = 5
    for q in question_bank.get_all():
        count = usage_counts.get(q.question_id, 0)
        ws.cell(row=row, column=1, value=q.question_no).border = thin_border
        ws.cell(row=row, column=1).alignment = center_align
        ws.cell(row=row, column=2, value=q.question_id).border = thin_border
        ws.cell(row=row, column=3, value=q.difficulty.capitalize()).border = thin_border
        ws.cell(row=row, column=4, value=count).border = thin_border
        ws.cell(row=row, column=4).alignment = center_align
        row += 1

    # ── Min/Max/Delta by Difficulty ──
    row += 1
    ws.cell(row=row, column=1, value="Min / Max / Delta by Difficulty").font = Font(bold=True, size=12)
    row += 1

    for col, h in enumerate(['Difficulty', 'Min', 'Max', 'Delta', 'Variance'], 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = header_font_white
        cell.fill = orange_fill
        cell.border = thin_border
    row += 1

    by_diff = defaultdict(list)
    for q in question_bank.get_all():
        by_diff[q.difficulty].append(usage_counts.get(q.question_id, 0))

    all_counts = list(usage_counts.values())

    # Track min/max per difficulty for overall calculation
    diff_stats = []

    for diff in ['hard', 'medium', 'easy']:
        counts = by_diff.get(diff, [])
        if counts:
            mn, mx = min(counts), max(counts)
            var = round(float(np.var(counts)), 4)
        else:
            mn = mx = 0
            var = 0.0

        diff_stats.append((mn, mx))

        ws.cell(row=row, column=1, value=diff.capitalize()).border = thin_border
        ws.cell(row=row, column=2, value=mn).border = thin_border
        ws.cell(row=row, column=3, value=mx).border = thin_border
        ws.cell(row=row, column=4, value=mx - mn).border = thin_border
        ws.cell(row=row, column=5, value=var).border = thin_border
        row += 1

    # Overall row: sum of min/max from each difficulty
    overall_min = sum(mn for mn, mx in diff_stats)
    overall_max = sum(mx for mn, mx in diff_stats)
    overall_delta = overall_max - overall_min

    ws.cell(row=row, column=1, value="OVERALL").border = thin_border
    ws.cell(row=row, column=1).font = Font(bold=True)
    ws.cell(row=row, column=2, value=overall_min).border = thin_border
    ws.cell(row=row, column=3, value=overall_max).border = thin_border
    ws.cell(row=row, column=4, value=overall_delta).border = thin_border
    ws.cell(row=row, column=5, value="-").border = thin_border

    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 12

    # ══════════════════════════════════════════════════════════════════════
    # Question Bank Sheet (embedded for Part 2 answer checking)
    # ══════════════════════════════════════════════════════════════════════
    ws = wb.create_sheet(title="Question_Bank")
    ws['A1'] = "Question Bank (Embedded for Answer Checking)"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:H1')

    qb_headers = ['question_no', 'question', 'option_a', 'option_b',
                   'option_c', 'option_d', 'answer', 'difficulty']
    qb_fill = PatternFill(start_color="8DB4E2", end_color="8DB4E2", fill_type="solid")
    for col, h in enumerate(qb_headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = header_font_white
        cell.fill = qb_fill
        cell.border = thin_border

    for q_idx, q in enumerate(question_bank.get_all()):
        row = q_idx + 4
        ws.cell(row=row, column=1, value=q.question_no).border = thin_border
        ws.cell(row=row, column=1).alignment = center_align
        ws.cell(row=row, column=2, value=q.question_text).border = thin_border
        ws.cell(row=row, column=2).alignment = wrap_align
        ws.cell(row=row, column=3, value=q.option_a).border = thin_border
        ws.cell(row=row, column=4, value=q.option_b).border = thin_border
        ws.cell(row=row, column=5, value=q.option_c).border = thin_border
        ws.cell(row=row, column=6, value=q.option_d).border = thin_border
        ws.cell(row=row, column=7, value=q.answer).border = thin_border
        ws.cell(row=row, column=7).alignment = center_align
        ws.cell(row=row, column=8, value=q.difficulty.capitalize()).border = thin_border

    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 50
    for ch in 'CDEF':
        ws.column_dimensions[ch].width = 20
    ws.column_dimensions['G'].width = 10
    ws.column_dimensions['H'].width = 12

    # ── Save ──────────────────────────────────────────────────────────────
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def _make_excel_bytes_from_dataframe(df: pd.DataFrame, sheet_name: str) -> bytes:
    """Serialize DataFrame into an in-memory Excel file."""
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)
    output.seek(0)
    return output.getvalue()


def _load_question_bank_from_question_papers(question_papers_path: str) -> FullQuestionBank:
    """Load embedded Question_Bank sheet from question_papers.xlsx."""
    required_cols = [
        "question_no",
        "question",
        "option_a",
        "option_b",
        "option_c",
        "option_d",
        "answer",
        "difficulty",
    ]
    normalized_required = set(required_cols)

    def _normalize_cols(df: pd.DataFrame) -> pd.DataFrame:
        out = df.copy()
        out.columns = [str(c).strip().lower().replace(" ", "_") for c in out.columns]
        return out

    temp_path = None
    try:
        try:
            # Support both formats:
            # 1) plain table with header on first row
            # 2) styled sheet with title row and header at row 3 (0-index header=2)
            question_bank_df = None
            for header_row in (0, 1, 2, 3, 4):
                candidate = pd.read_excel(question_papers_path, sheet_name="Question_Bank", header=header_row)
                candidate = _normalize_cols(candidate)
                if normalized_required.issubset(set(candidate.columns)):
                    question_bank_df = candidate[required_cols].copy()
                    question_bank_df = question_bank_df.dropna(how="all")
                    break

            if question_bank_df is None:
                raise ValueError(
                    f"Missing required columns: {required_cols}"
                )
        except ValueError as exc:
            if "Worksheet named 'Question_Bank' not found" in str(exc):
                raise ValueError(
                    "Question_Bank sheet not found in question papers. "
                    "Regenerate papers with the latest app/CLI."
                ) from exc
            raise ValueError(
                "Question_Bank sheet is present but not in expected format. "
                "Regenerate papers with the latest app/CLI."
            ) from exc
        except Exception as exc:
            raise ValueError(
                "Question_Bank sheet not found in question papers. "
                "Regenerate papers with the latest app/CLI."
            ) from exc

        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx", prefix="embedded_qb_") as tmp:
            temp_path = tmp.name

        with pd.ExcelWriter(temp_path, engine="openpyxl") as writer:
            question_bank_df.to_excel(writer, index=False)

        return load_question_bank(temp_path)
    finally:
        if temp_path and os.path.exists(temp_path):
            os.remove(temp_path)
